"""
=====================================================================
  Program Name  : scanFileSystem.py
  Author        : Manuel Figallo
  Purpose       : General-purpose file-system scanner and text-extraction
                  utility. It crawls one or more directory roots, captures
                  file-system metadata for every matching file, extracts
                  caller-specified keywords together with surrounding
                  context, optionally filters by a date range, and (via an
                  opt-in metric profile) parses structured performance
                  metrics from log files. The flagship profile extracts SAS
                  per-step "real time" and "cpu time", but the same engine
                  generalizes to any keyword sweep or log-metric use case.
  Version       : 1.3.5
  Created       : 2026-08-20
  Last Modified : 2026-08-25

  Dependencies:
    CSV output requires nothing beyond the Python standard library. XLSX
    output requires openpyxl (or xlsxwriter); without either, the scan falls
    back to CSV with a warning. pandas/numpy are NOT used.

  Description:
    Runs unattended on a schedule (Windows Task Scheduler) or, optionally,
    from a SAS SYSTASK wrapper. Fully parameterized at the top of the file
    with CLI overrides; never prompts interactively. Validates the required
    parameter at startup and exits non-zero with a clear ERROR if it is
    missing. Emits a Files grain (one row per file), plus a StepDetail grain
    (second .xlsx sheet or companion "_StepDetail.csv") when a metric
    profile is active.

  Configuration:
    Edit ONLY the CONFIG block near the top of this file (or pass the
    equivalent --kebab-case CLI flags) to change roots, extensions, folder
    exclusions, output path, keywords, date range, or the metric profile.

  Usage:
    Unattended (Task Scheduler / cmd):
      python scanFileSystem.py --input-folder-root "\\\\srv\\logs" --output-file-path "C:\\Logs\\scan.xlsx"
    From SAS (optional): via the %scanFileSystem() SYSTASK wrapper.

  Input Parameters (required first):
    input_folder_root      (REQUIRED, list[str]) - root path(s) to search;
                             single string or array. Empty/None -> ERROR + exit.
    output_file_path       (optional, str) - .csv or .xlsx by extension;
                             a directory auto-names a timestamped .csv inside
                             it. When omitted entirely, writes
                             scan_YYYYMMDD_HHMMSS.csv to the current directory.
    file_extensions        (list[str], default ["log","txt","sas"]) -
                             extensions to include; case-insensitive.
    include_subdirectories (bool, default True) - recurse when True.
    folder_exclusion_list  (list[str], default []) - folder names/tokens to
                             exclude (e.g. ["Old","Test"]); empty means nothing
                             excluded. Matches an ancestor directory segment
                             (case-insensitive) or a full-path prefix.
    file_exclusion_list    (list[str], default []) - prefixes/tokens stripped
                             from the filename to derive program_name.
    extract_keyword        (list[str], default []) - keywords to extract from
                             any text; each yields a matched line, a +/-3 line
                             context window, and a match count.
    date_from              (str, default None) - inclusive lower bound
                             (YYYY-MM-DD / ISO) for the date-range filter.
    date_to                (str, default None) - inclusive upper bound.
    date_field             (str, default "modified") - created/modified/accessed.
    metric_profile         (str, default "none") - "none" disables StepDetail;
                             "sas_log" extracts per-step real/cpu time.

  Output:
    Files grain (one row per file): program_name, log_file_name, full_path,
    directory, extension, file_size_bytes, created/modified/accessed_time,
    step_count, total_real_time_sec, total_cpu_time_sec,
    max_step_real_time_sec, max_step_label, error_count, warning_count,
    kw_<K>_line / kw_<K>_context / kw_<K>_count per keyword, parse_status,
    scanned_at.
    StepDetail grain (when a profile is active): full_path, program_name,
    step_index, step_label, real_time_sec, cpu_time_sec - as a second
    "StepDetail" sheet (.xlsx) or a companion "<stem>_StepDetail.csv" (.csv).

  Change Log:
    v1.0.0 - Initial release.
    v1.1.0 - Added StepDetail grain; generalized with pluggable metric
             profiles; SAS wrapper made optional.
    v1.2.0 - Built via Claude Code with synthetic fixtures and self-tests.
    v1.3.0 - folder_exclusion_list replaces the CSV; required-param
             validation; metric_profile defaults to "none"; optional
             date-range filter.
    v1.3.1 - folder_exclusion_list now defaults to empty (nothing excluded
             unless specified).
    v1.3.2 - output_file_path is now optional; when omitted the scan writes
             scan_YYYYMMDD_HHMMSS.csv to the current directory.
    v1.3.3 - Dropped the pandas/numpy dependency: CSV output is now pure
             standard library and XLSX is written directly via openpyxl (or
             xlsxwriter). Output path is resolved before crawling so a bad
             path fails fast. Control characters Excel rejects are stripped.
    v1.3.5 - Version aligned across the Python, PowerShell and SAS
             components so a stale copy is obvious in the logs.
=====================================================================
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

__version__ = "1.3.5"

# =====================================================================
# CONFIG -- edit ONLY this block (or pass the equivalent CLI flags).
# =====================================================================

input_folder_root: List[str] = [
    r"\\a70admed.com\R1\CGS\APPS\SAS\UNIT\SAS_G\SAS\Manuel\data\logs\UNIT"
]
output_file_path: Optional[str] = None   # optional; None -> scan_YYYYMMDD_HHMMSS.csv
file_extensions: List[str] = ["log", "txt", "sas"]
include_subdirectories: bool = True
folder_exclusion_list: List[str] = []          # v1.3.1: empty = nothing excluded
file_exclusion_list: List[str] = []
extract_keyword: List[str] = []
date_from: Optional[str] = None
date_to: Optional[str] = None
date_field: str = "modified"                    # created | modified | accessed
metric_profile: str = "none"                    # none | sas_log

# =====================================================================
# Constants
# =====================================================================

CONTEXT_LINES = 3                               # +/- N lines around a keyword hit
VALID_DATE_FIELDS = ("created", "modified", "accessed")

#: Auto-generated output name when output_file_path is not supplied:
#: "scan" + _YYYYMMDD_HHMMSS + ".csv", written to the current directory.
DEFAULT_OUTPUT_PREFIX = "scan"
DEFAULT_OUTPUT_SUFFIX = ".csv"
TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"

EXIT_OK = 0
EXIT_CONFIG_ERROR = 2                           # bad/missing params, bad dates
EXIT_IO_ERROR = 3                               # unreachable roots, unwritable output

FILES_SHEET = "Files"
STEPDETAIL_SHEET = "StepDetail"

FILES_BASE_COLUMNS = [
    "program_name", "log_file_name", "full_path", "directory", "extension",
    "file_size_bytes", "created_time", "modified_time", "accessed_time",
    "step_count", "total_real_time_sec", "total_cpu_time_sec",
    "max_step_real_time_sec", "max_step_label", "error_count", "warning_count",
]
FILES_TAIL_COLUMNS = ["parse_status", "scanned_at"]
STEPDETAIL_COLUMNS = [
    "full_path", "program_name", "step_index", "step_label",
    "real_time_sec", "cpu_time_sec",
]


# =====================================================================
# Logging (stderr only -- never interactive)
# =====================================================================

def _log(level: str, message: str) -> None:
    """Write a timestamped log line to stderr. Never prompts, never blocks."""
    stamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{stamp} {level:<7} {message}", file=sys.stderr, flush=True)


def log_info(msg: str) -> None:
    _log("INFO", msg)


def log_warn(msg: str) -> None:
    _log("WARNING", msg)


def log_error(msg: str) -> None:
    _log("ERROR", msg)


# =====================================================================
# Metric profiles (pluggable; regex/config-driven)
#
# A profile turns file text into StepDetail rows plus a small dict of
# file-level counters. Adding a profile means adding an entry to
# METRIC_PROFILES -- no changes to the crawl or output code.
# =====================================================================

def parse_duration(raw: str) -> Optional[float]:
    """Parse a duration into seconds.

    Handles plain seconds ("0.05", "1.20 seconds") and clock forms
    ("1:03.05" = mm:ss, "1:00:30.00" = hh:mm:ss).
    """
    if raw is None:
        return None
    text = str(raw).strip().lower().replace("seconds", "").replace("second", "").strip()
    if not text:
        return None
    m = re.match(r"^(\d+):(\d{1,2}):(\d{1,2}(?:\.\d+)?)$", text)   # hh:mm:ss
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    m = re.match(r"^(\d+):(\d{1,2}(?:\.\d+)?)$", text)             # mm:ss
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))
    m = re.match(r"^(\d+(?:\.\d+)?)$", text)                       # plain seconds
    if m:
        return float(m.group(1))
    return None


class MetricProfile:
    """Base profile: parses nothing. Used when metric_profile="none"."""

    name = "none"
    active = False

    def parse(self, lines: Sequence[str]) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
        """Return (step rows, file-level counters)."""
        return [], {"error_count": 0, "warning_count": 0}


class RegexStepProfile(MetricProfile):
    """Config-driven profile: a step-header regex plus per-metric regexes.

    step_pattern must expose a named group ``label``. For each header match the
    next ``lookahead`` lines are searched for each metric pattern, which must
    expose a named group ``value``. Counter patterns are counted line-wise
    across the whole file.
    """

    active = True

    def __init__(
        self,
        name: str,
        step_pattern: str,
        metric_patterns: Dict[str, str],
        counter_patterns: Optional[Dict[str, str]] = None,
        lookahead: int = 6,
    ) -> None:
        self.name = name
        self._step_re = re.compile(step_pattern, re.IGNORECASE)
        self._metric_res = {
            key: re.compile(pat, re.IGNORECASE) for key, pat in metric_patterns.items()
        }
        self._counter_res = {
            key: re.compile(pat, re.IGNORECASE)
            for key, pat in (counter_patterns or {}).items()
        }
        self._lookahead = lookahead

    def parse(self, lines: Sequence[str]) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
        steps: List[Dict[str, Any]] = []
        for idx, line in enumerate(lines):
            header = self._step_re.search(line)
            if not header:
                continue
            step: Dict[str, Any] = {
                "step_index": len(steps) + 1,
                "step_label": (header.groupdict().get("label") or "").strip(),
            }
            window = lines[idx + 1: idx + 1 + self._lookahead]
            for key, regex in self._metric_res.items():
                step[key] = None
                for candidate in window:
                    hit = regex.search(candidate)
                    if hit:
                        step[key] = parse_duration(hit.groupdict().get("value", ""))
                        break
            steps.append(step)

        counters = {key: 0 for key in self._counter_res}
        for line in lines:
            for key, regex in self._counter_res.items():
                if regex.search(line):
                    counters[key] += 1
        counters.setdefault("error_count", 0)
        counters.setdefault("warning_count", 0)
        return steps, counters


#: Profile registry. Add a new profile here -- nothing else changes.
METRIC_PROFILES: Dict[str, MetricProfile] = {
    "none": MetricProfile(),
    "sas_log": RegexStepProfile(
        name="sas_log",
        # NOTE: DATA statement used (Total process time):
        step_pattern=r"^NOTE:\s+(?P<label>.+?)\s+used\s+\(Total process time\)",
        metric_patterns={
            "real_time_sec": r"^\s*real time\s+(?P<value>[0-9:.]+)",
            "cpu_time_sec": r"^\s*cpu time\s+(?P<value>[0-9:.]+)",
        },
        counter_patterns={
            "error_count": r"^\s*ERROR[: ]",
            "warning_count": r"^\s*WARNING[: ]",
        },
    ),
}


def aggregate_steps(steps: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    """Roll StepDetail rows up to the Files grain."""
    reals = [s.get("real_time_sec") for s in steps if s.get("real_time_sec") is not None]
    cpus = [s.get("cpu_time_sec") for s in steps if s.get("cpu_time_sec") is not None]
    max_label, max_real = "", 0.0
    for step in steps:
        value = step.get("real_time_sec")
        if value is not None and value > max_real:
            max_real, max_label = value, step.get("step_label", "")
    return {
        "step_count": len(steps),
        "total_real_time_sec": round(sum(reals), 6) if reals else 0.0,
        "total_cpu_time_sec": round(sum(cpus), 6) if cpus else 0.0,
        "max_step_real_time_sec": max_real,
        "max_step_label": max_label,
    }


# =====================================================================
# Keyword extraction (generic -- any text, any keyword)
# =====================================================================

def keyword_slug(keyword: str) -> str:
    """Turn a keyword into a safe column-name fragment ('.accdb' -> 'accdb')."""
    slug = re.sub(r"[^0-9a-zA-Z]+", "_", str(keyword)).strip("_").lower()
    return slug or "kw"


def keyword_columns(keywords: Sequence[str]) -> List[str]:
    """Build the ordered kw_* column list, de-duplicating collided slugs."""
    columns: List[str] = []
    for slug in _unique_slugs(keywords):
        columns.extend([f"kw_{slug}_line", f"kw_{slug}_context", f"kw_{slug}_count"])
    return columns


def _unique_slugs(keywords: Sequence[str]) -> List[str]:
    slugs, seen = [], {}
    for keyword in keywords:
        base = keyword_slug(keyword)
        seen[base] = seen.get(base, 0) + 1
        slugs.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return slugs


def extract_keywords(lines: Sequence[str], keywords: Sequence[str]) -> Dict[str, Any]:
    """First matched line, a +/-CONTEXT_LINES window, and a match count each."""
    out: Dict[str, Any] = {}
    slugs = _unique_slugs(keywords)
    for keyword, slug in zip(keywords, slugs):
        needle = str(keyword).lower()
        first_line, context, count = "", "", 0
        for idx, line in enumerate(lines):
            if needle in line.lower():
                count += 1
                if not first_line:
                    first_line = line.strip()
                    lo = max(0, idx - CONTEXT_LINES)
                    hi = min(len(lines), idx + CONTEXT_LINES + 1)
                    context = "\n".join(l.rstrip("\n") for l in lines[lo:hi])
        out[f"kw_{slug}_line"] = first_line
        out[f"kw_{slug}_context"] = context
        out[f"kw_{slug}_count"] = count
    return out


# =====================================================================
# Crawl / filter / metadata (format-agnostic)
# =====================================================================

def normalize_path_text(text: str) -> str:
    """Lowercase, forward-slash, trailing-slash-safe form for comparisons."""
    return str(text).replace("\\", "/").rstrip("/").lower()


def is_folder_excluded(path: Path, exclusions: Sequence[str]) -> bool:
    """True when an ancestor directory segment matches, or a full-path prefix does.

    Segment matching is exact per segment (case-insensitive), so excluding
    "Old" does NOT exclude a sibling "Older".
    """
    if not exclusions:
        return False
    parent_segments = {str(part).lower() for part in path.parent.parts}
    normalized_path = normalize_path_text(path)
    for raw in exclusions:
        token = str(raw).strip()
        if not token:
            continue
        if token.replace("\\", "/").rstrip("/").lower() in parent_segments:
            return True
        prefix = normalize_path_text(token)
        if prefix and (normalized_path == prefix or normalized_path.startswith(prefix + "/")):
            return True
    return False


def normalize_extensions(extensions: Sequence[str]) -> set:
    """'log' / '.LOG' -> 'log'. Empty list means 'accept everything'."""
    return {str(e).strip().lstrip(".").lower() for e in extensions if str(e).strip()}


def parse_boundary(raw: Optional[str], *, end_of_day: bool) -> Optional[dt.datetime]:
    """Parse a YYYY-MM-DD or ISO datetime bound; raise ValueError if malformed.

    A bare date used as an upper bound extends to the end of that day so the
    bound is genuinely inclusive.
    """
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip()
    try:
        value = dt.datetime.fromisoformat(text)
    except ValueError:
        raise ValueError(f"unparseable date {raw!r}; expected YYYY-MM-DD or ISO datetime")
    if end_of_day and len(text) == 10:      # bare date as an upper bound
        value = value.replace(hour=23, minute=59, second=59, microsecond=999999)
    return value


def stat_times(stat_result: os.stat_result) -> Dict[str, dt.datetime]:
    return {
        "created": dt.datetime.fromtimestamp(getattr(stat_result, "st_ctime", 0)),
        "modified": dt.datetime.fromtimestamp(stat_result.st_mtime),
        "accessed": dt.datetime.fromtimestamp(stat_result.st_atime),
    }


def derive_program_name(filename: str, exclusions: Sequence[str]) -> str:
    """Filename minus extension, minus any configured prefix/token."""
    stem = Path(filename).stem
    for raw in exclusions:
        token = str(raw).strip()
        if not token:
            continue
        if stem.lower().startswith(token.lower()):
            stem = stem[len(token):]
        else:
            stem = re.sub(re.escape(token), "", stem, flags=re.IGNORECASE)
    return stem.strip(" _-.") or Path(filename).stem


def read_text(path: Path) -> List[str]:
    """utf-8, falling back to latin-1, then utf-8 with replacement."""
    data = path.read_bytes()
    for encoding, errors in (("utf-8", "strict"), ("latin-1", "strict"), ("utf-8", "replace")):
        try:
            return data.decode(encoding, errors).splitlines()
        except (UnicodeDecodeError, LookupError):
            continue
    return []


def iter_candidate_files(roots: Sequence[str], recurse: bool) -> Tuple[List[Path], int]:
    """Yield every file under the roots. Missing roots are logged and skipped."""
    found: List[Path] = []
    reachable = 0
    for raw_root in roots:
        root = Path(str(raw_root).strip())
        if not root.exists():
            log_warn(f"root not found, skipping: {root}")
            continue
        if not root.is_dir():
            log_warn(f"root is not a directory, skipping: {root}")
            continue
        reachable += 1
        try:
            if recurse:
                for dirpath, _dirnames, filenames in os.walk(root, onerror=_walk_error):
                    for name in filenames:
                        found.append(Path(dirpath) / name)
            else:
                found.extend(child for child in root.iterdir() if child.is_file())
        except PermissionError as exc:
            log_warn(f"permission denied while walking {root}: {exc}")
    return found, reachable


def _walk_error(exc: OSError) -> None:
    log_warn(f"cannot descend into {getattr(exc, 'filename', '?')}: {exc}")


# =====================================================================
# Scan
# =====================================================================

def scan(
    roots: Sequence[str],
    *,
    extensions: Sequence[str],
    recurse: bool,
    folder_exclusions: Sequence[str],
    file_exclusions: Sequence[str],
    keywords: Sequence[str],
    profile: MetricProfile,
    date_lo: Optional[dt.datetime] = None,
    date_hi: Optional[dt.datetime] = None,
    which_date: str = "modified",
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], int]:
    """Crawl, filter, extract. Returns (files_rows, stepdetail_rows, reachable_roots)."""
    wanted = normalize_extensions(extensions)
    scanned_at = dt.datetime.now().isoformat(timespec="seconds")
    files_rows: List[Dict[str, Any]] = []
    step_rows: List[Dict[str, Any]] = []

    candidates, reachable = iter_candidate_files(roots, recurse)
    log_info(f"discovered {len(candidates)} file(s) across {reachable} reachable root(s)")

    for path in sorted(candidates):
        extension = path.suffix.lstrip(".").lower()
        if wanted and extension not in wanted:
            continue
        if is_folder_excluded(path, folder_exclusions):
            continue

        program_name = derive_program_name(path.name, file_exclusions)
        row: Dict[str, Any] = {
            "program_name": program_name,
            "log_file_name": path.name,
            "full_path": str(path),
            "directory": str(path.parent),
            "extension": extension,
            "file_size_bytes": 0,
            "created_time": "", "modified_time": "", "accessed_time": "",
            "step_count": 0, "total_real_time_sec": 0.0, "total_cpu_time_sec": 0.0,
            "max_step_real_time_sec": 0.0, "max_step_label": "",
            "error_count": 0, "warning_count": 0,
            "parse_status": "OK",
            "scanned_at": scanned_at,
        }
        for column in keyword_columns(keywords):
            row[column] = 0 if column.endswith("_count") else ""

        # --- metadata + date filter (a stat failure still emits a row) ---
        try:
            info = path.stat()
            times = stat_times(info)
            if date_lo is not None and times[which_date] < date_lo:
                continue
            if date_hi is not None and times[which_date] > date_hi:
                continue
            row["file_size_bytes"] = info.st_size
            row["created_time"] = times["created"].isoformat(timespec="seconds")
            row["modified_time"] = times["modified"].isoformat(timespec="seconds")
            row["accessed_time"] = times["accessed"].isoformat(timespec="seconds")
        except OSError as exc:
            row["parse_status"] = f"stat error: {exc.__class__.__name__}: {exc}"
            log_warn(f"cannot stat {path}: {exc}")
            files_rows.append(row)
            continue

        # --- content: keywords + metric profile ---
        try:
            lines = read_text(path)
            if keywords:
                row.update(extract_keywords(lines, keywords))
            steps, counters = profile.parse(lines)
            row["error_count"] = counters.get("error_count", 0)
            row["warning_count"] = counters.get("warning_count", 0)
            if profile.active:
                row.update(aggregate_steps(steps))
                for step in steps:
                    step_rows.append({
                        "full_path": str(path),
                        "program_name": program_name,
                        "step_index": step.get("step_index"),
                        "step_label": step.get("step_label", ""),
                        "real_time_sec": step.get("real_time_sec"),
                        "cpu_time_sec": step.get("cpu_time_sec"),
                    })
        except (OSError, UnicodeError, ValueError) as exc:
            row["parse_status"] = f"read error: {exc.__class__.__name__}: {exc}"
            log_warn(f"cannot read {path}: {exc}")

        files_rows.append(row)

    return files_rows, step_rows, reachable


# =====================================================================
# Output
# =====================================================================

def default_output_name(directory: Optional[Path] = None) -> Path:
    """Build the auto-generated output name: scan_YYYYMMDD_HHMMSS.csv.

    Used when output_file_path is not supplied (writes to the current
    directory) and when the caller supplies a directory instead of a file.
    """
    stamp = dt.datetime.now().strftime(TIMESTAMP_FORMAT)
    filename = f"{DEFAULT_OUTPUT_PREFIX}_{stamp}{DEFAULT_OUTPUT_SUFFIX}"
    return (directory / filename) if directory is not None else Path(filename)


def resolve_output_path(raw: str) -> Path:
    """Resolve the output target; a directory auto-names a timestamped .csv."""
    path = Path(str(raw).strip())
    looks_like_dir = (
        (path.exists() and path.is_dir())
        or str(raw).rstrip().endswith(("/", "\\"))
        or path.suffix.lower() not in (".csv", ".xlsx")
    )
    if looks_like_dir:
        path.mkdir(parents=True, exist_ok=True)
        generated = default_output_name(path)
        log_info(f"output path is a directory; writing {generated}")
        return generated
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


#: Control characters Excel refuses inside a cell (openpyxl raises on these).
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def excel_safe(value: Any) -> Any:
    """Strip control characters Excel rejects; pass non-strings through."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def first_available_excel_engine() -> Optional[str]:
    """openpyxl -> xlsxwriter -> None (caller falls back to CSV)."""
    for engine in ("openpyxl", "xlsxwriter"):
        try:
            __import__(engine)
            return engine
        except ImportError:
            continue
    return None


def _write_csv(rows: List[Dict[str, Any]], columns: List[str], target: Path) -> None:
    """Write rows as CSV using only the standard library (no index column)."""
    with open(target, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def _write_xlsx(sheets: List[Tuple[str, List[Dict[str, Any]], List[str]]],
                target: Path, engine: str) -> None:
    """Write one or more sheets directly via openpyxl or xlsxwriter."""
    if engine == "openpyxl":
        from openpyxl import Workbook
        workbook = Workbook()
        workbook.remove(workbook.active)            # drop the default sheet
        for title, rows, columns in sheets:
            sheet = workbook.create_sheet(title)
            sheet.append(columns)
            for row in rows:
                sheet.append([excel_safe(row.get(column, "")) for column in columns])
        workbook.save(target)
        return

    import xlsxwriter
    workbook = xlsxwriter.Workbook(str(target), {"constant_memory": True})
    try:
        for title, rows, columns in sheets:
            sheet = workbook.add_worksheet(title)
            for index, column in enumerate(columns):
                sheet.write(0, index, column)
            for row_index, row in enumerate(rows, start=1):
                for index, column in enumerate(columns):
                    sheet.write(row_index, index, excel_safe(row.get(column, "")))
    finally:
        workbook.close()


def write_output(
    files_rows: List[Dict[str, Any]],
    step_rows: List[Dict[str, Any]],
    target: Path,
    keywords: Sequence[str],
    profile_active: bool,
) -> List[Path]:
    """Write the Files grain (+ StepDetail when a profile is active).

    CSV output is pure standard library. XLSX uses openpyxl or xlsxwriter
    directly -- no pandas/numpy anywhere on the runtime path.
    """
    columns = FILES_BASE_COLUMNS + keyword_columns(keywords) + FILES_TAIL_COLUMNS
    written: List[Path] = []

    if target.suffix.lower() == ".xlsx":
        engine = first_available_excel_engine()
        if engine is None:
            log_warn("no Excel engine (openpyxl/xlsxwriter) available; falling back to CSV")
            target = target.with_suffix(".csv")
        else:
            sheets = [(FILES_SHEET, files_rows, columns)]
            if profile_active:
                sheets.append((STEPDETAIL_SHEET, step_rows, STEPDETAIL_COLUMNS))
            _write_xlsx(sheets, target, engine)
            written.append(target)
            log_info(f"wrote {len(files_rows)} Files row(s) to {target} [{engine}]")
            if profile_active:
                log_info(f"wrote {len(step_rows)} StepDetail row(s) to sheet '{STEPDETAIL_SHEET}'")
            return written

    _write_csv(files_rows, columns, target)
    written.append(target)
    log_info(f"wrote {len(files_rows)} Files row(s) to {target}")
    if profile_active:
        companion = target.with_name(f"{target.stem}_StepDetail.csv")
        _write_csv(step_rows, STEPDETAIL_COLUMNS, companion)
        written.append(companion)
        log_info(f"wrote {len(step_rows)} StepDetail row(s) to companion {companion}")
    return written


# =====================================================================
# CLI / validation
# =====================================================================

def _as_list(value: Any) -> List[str]:
    """Accept a single string or an array; split semicolon-delimited strings."""
    if value is None:
        return []
    if isinstance(value, str):
        return [part.strip() for part in value.split(";") if part.strip()]
    out: List[str] = []
    for item in value:
        out.extend(_as_list(item) if isinstance(item, str) else [str(item)])
    return out


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="scanFileSystem.py",
        description="General-purpose file-system scanner and text-extraction utility.",
    )
    parser.add_argument("--input-folder-root", nargs="+", default=None,
                        help="One or more root paths to scan (REQUIRED).")
    parser.add_argument("--output-file-path", default=None,
                        help=".csv or .xlsx by extension; a directory auto-names a "
                             "timestamped .csv inside it. Optional -- omit to write "
                             "scan_YYYYMMDD_HHMMSS.csv in the current directory.")
    parser.add_argument("--file-extensions", nargs="+", default=None,
                        help='Extensions to include (default: log txt sas).')
    parser.add_argument("--include-subdirectories", action=argparse.BooleanOptionalAction,
                        default=None, help="Recurse into subdirectories (default: on).")
    parser.add_argument("--folder-exclusion-list", nargs="*", default=None,
                        help='Folder names/tokens to exclude, e.g. Old Test (default: none).')
    parser.add_argument("--file-exclusion-list", nargs="*", default=None,
                        help="Prefixes/tokens stripped from the filename for program_name.")
    parser.add_argument("--extract-keyword", nargs="*", default=None,
                        help="Keywords to extract with +/-3 lines of context.")
    parser.add_argument("--date-from", default=None, help="Inclusive lower bound (YYYY-MM-DD/ISO).")
    parser.add_argument("--date-to", default=None, help="Inclusive upper bound (YYYY-MM-DD/ISO).")
    parser.add_argument("--date-field", default=None, choices=list(VALID_DATE_FIELDS),
                        help="Which timestamp the date range filters on (default: modified).")
    parser.add_argument("--metric-profile", default=None,
                        help=f"Metric profile: {' | '.join(METRIC_PROFILES)} (default: none).")
    parser.add_argument("--version", action="version", version=f"scanFileSystem {__version__}")
    return parser


def resolve_settings(args: argparse.Namespace) -> Dict[str, Any]:
    """CLI flags override the CONFIG block; CONFIG supplies the defaults."""
    return {
        "roots": _as_list(args.input_folder_root if args.input_folder_root is not None
                          else input_folder_root),
        "output": (args.output_file_path if args.output_file_path is not None
                   else output_file_path),
        "extensions": _as_list(args.file_extensions if args.file_extensions is not None
                               else file_extensions),
        "recurse": (include_subdirectories if args.include_subdirectories is None
                    else args.include_subdirectories),
        "folder_exclusions": _as_list(args.folder_exclusion_list
                                      if args.folder_exclusion_list is not None
                                      else folder_exclusion_list),
        "file_exclusions": _as_list(args.file_exclusion_list
                                    if args.file_exclusion_list is not None
                                    else file_exclusion_list),
        "keywords": _as_list(args.extract_keyword if args.extract_keyword is not None
                             else extract_keyword),
        "date_from": args.date_from if args.date_from is not None else date_from,
        "date_to": args.date_to if args.date_to is not None else date_to,
        "date_field": (args.date_field if args.date_field is not None else date_field),
        "metric_profile": (args.metric_profile if args.metric_profile is not None
                           else metric_profile),
    }


def validate(settings: Dict[str, Any]) -> Optional[str]:
    """Return an error message for the first invalid setting, else None."""
    if not settings["roots"]:
        return ("required parameter 'input_folder_root' is missing or empty; "
                "pass --input-folder-root or set it in the CONFIG block")
    # output_file_path is OPTIONAL (v1.3.2): when absent, main() auto-names
    # scan_YYYYMMDD_HHMMSS.csv in the current directory.
    if settings["metric_profile"] not in METRIC_PROFILES:
        return (f"unknown metric_profile {settings['metric_profile']!r}; "
                f"expected one of: {', '.join(METRIC_PROFILES)}")
    if settings["date_field"] not in VALID_DATE_FIELDS:
        return (f"unknown date_field {settings['date_field']!r}; "
                f"expected one of: {', '.join(VALID_DATE_FIELDS)}")
    return None


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    settings = resolve_settings(args)

    problem = validate(settings)
    if problem:
        log_error(problem)
        return EXIT_CONFIG_ERROR

    try:
        date_lo = parse_boundary(settings["date_from"], end_of_day=False)
        date_hi = parse_boundary(settings["date_to"], end_of_day=True)
    except ValueError as exc:
        log_error(str(exc))
        return EXIT_CONFIG_ERROR
    if date_lo and date_hi and date_lo > date_hi:
        log_error(f"date_from ({settings['date_from']}) is after date_to ({settings['date_to']})")
        return EXIT_CONFIG_ERROR

    profile = METRIC_PROFILES[settings["metric_profile"]]
    log_info(f"scanFileSystem {__version__} starting; profile={profile.name}; "
             f"roots={len(settings['roots'])}")

    # Resolve (and create the parent of) the output target BEFORE crawling, so a
    # bad path fails in seconds instead of after a long network scan.
    try:
        if settings["output"] and str(settings["output"]).strip():
            target = resolve_output_path(settings["output"])
        else:
            target = default_output_name()
            log_info(f"output_file_path not supplied; writing {target}")
    except OSError as exc:
        log_error(f"cannot prepare output path: {exc}")
        return EXIT_IO_ERROR

    try:
        files_rows, step_rows, reachable = scan(
            settings["roots"],
            extensions=settings["extensions"],
            recurse=settings["recurse"],
            folder_exclusions=settings["folder_exclusions"],
            file_exclusions=settings["file_exclusions"],
            keywords=settings["keywords"],
            profile=profile,
            date_lo=date_lo,
            date_hi=date_hi,
            which_date=settings["date_field"],
        )
    except OSError as exc:
        log_error(f"fatal I/O error while scanning: {exc}")
        return EXIT_IO_ERROR

    if reachable == 0:
        log_error("none of the supplied input_folder_root path(s) are reachable")
        return EXIT_IO_ERROR

    try:
        written = write_output(files_rows, step_rows, target,
                               settings["keywords"], profile.active)
    except (OSError, ImportError) as exc:
        log_error(f"cannot write output: {exc}")
        return EXIT_IO_ERROR

    log_info(f"done; {len(files_rows)} file row(s), {len(step_rows)} step row(s), "
             f"{len(written)} output file(s)")
    return EXIT_OK


if __name__ == "__main__":
    sys.exit(main())
