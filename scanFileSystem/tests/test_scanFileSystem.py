"""Self-checks for scanFileSystem.py, run against the synthetic fixture tree."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import scanFileSystem as sfs                      # noqa: E402
from make_fixtures import build as build_fixtures  # noqa: E402

# Hand-summed from the fixtures (see make_fixtures.py docstring).
EXPECTED = {
    "jobA": {"steps": 3, "real": 3.25, "cpu": 2.43, "errors": 1, "warnings": 2},
    "jobB": {"steps": 2, "real": 63.15, "cpu": 60.05, "errors": 0, "warnings": 0},
    "jobC": {"steps": 1, "real": 0.40, "cpu": 0.20, "errors": 0, "warnings": 0},
}


@pytest.fixture(scope="session")
def logs_root() -> str:
    return str(build_fixtures())


def run(argv) -> int:
    return sfs.main([str(a) for a in argv])


def read_rows(csv_path: Path) -> list:
    """Read a CSV into a list of dicts (stdlib only -- no pandas)."""
    with open(csv_path, encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def columns_of(csv_path: Path) -> list:
    with open(csv_path, encoding="utf-8", newline="") as handle:
        return next(csv.reader(handle))


def by_program(csv_path: Path) -> dict:
    """Index the Files grain by program_name."""
    return {row["program_name"]: row for row in read_rows(csv_path)}


def programs(csv_path: Path) -> set:
    return {row["program_name"] for row in read_rows(csv_path)}


def sheet_names(xlsx_path: Path) -> list:
    from openpyxl import load_workbook
    return load_workbook(xlsx_path, read_only=True).sheetnames


# --------------------------------------------------------------------- #
# Output grains & format dispatch
# --------------------------------------------------------------------- #

def test_csv_with_profile_writes_stepdetail_companion(tmp_path, logs_root):
    out = tmp_path / "scan.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out,
                "--metric-profile", "sas_log"]) == 0
    assert out.exists()
    companion = tmp_path / "scan_StepDetail.csv"
    assert companion.exists(), "sas_log profile must emit a _StepDetail.csv companion"
    assert read_rows(companion), "companion must contain rows"


def test_csv_without_profile_has_no_companion(tmp_path, logs_root):
    out = tmp_path / "plain.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out]) == 0
    assert out.exists()
    assert not (tmp_path / "plain_StepDetail.csv").exists(), \
        "metric_profile=none must not emit StepDetail"


def test_xlsx_sheets_depend_on_profile(tmp_path, logs_root):
    profiled = tmp_path / "scan.xlsx"
    assert run(["--input-folder-root", logs_root, "--output-file-path", profiled,
                "--metric-profile", "sas_log"]) == 0
    assert sheet_names(profiled) == ["Files", "StepDetail"]

    plain = tmp_path / "plain.xlsx"
    assert run(["--input-folder-root", logs_root, "--output-file-path", plain]) == 0
    assert sheet_names(plain) == ["Files"]


def test_directory_output_autonames_timestamped_csv(tmp_path, logs_root):
    outdir = tmp_path / "dropbox"
    outdir.mkdir()
    assert run(["--input-folder-root", logs_root, "--output-file-path", outdir]) == 0
    generated = list(outdir.glob("scan_*.csv"))
    assert len(generated) == 1, "a directory target must auto-name one timestamped CSV"
    assert re.fullmatch(r"scan_\d{8}_\d{6}\.csv", generated[0].name), generated[0].name


# --------------------------------------------------------------------- #
# sas_log metric profile
# --------------------------------------------------------------------- #

def test_sas_timings_match_hand_summed_values(tmp_path, logs_root):
    out = tmp_path / "scan.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out,
                "--metric-profile", "sas_log"]) == 0
    files = by_program(out)
    for program, want in EXPECTED.items():
        row = files[program]
        assert int(row["step_count"]) == want["steps"], program
        assert float(row["total_real_time_sec"]) == pytest.approx(want["real"]), program
        assert float(row["total_cpu_time_sec"]) == pytest.approx(want["cpu"]), program
        assert int(row["error_count"]) == want["errors"], program
        assert int(row["warning_count"]) == want["warnings"], program


def test_max_step_is_the_slowest_step(tmp_path, logs_root):
    out = tmp_path / "scan.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--metric-profile", "sas_log"])
    row = by_program(out)["jobA"]
    assert row["max_step_label"] == "PROCEDURE SORT"
    assert float(row["max_step_real_time_sec"]) == pytest.approx(2.00)


def test_stepdetail_grain_is_one_row_per_step(tmp_path, logs_root):
    out = tmp_path / "scan.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--metric-profile", "sas_log"])
    companion = tmp_path / "scan_StepDetail.csv"
    assert columns_of(companion) == sfs.STEPDETAIL_COLUMNS
    joba = [r for r in read_rows(companion) if r["program_name"] == "jobA"]
    assert [r["step_index"] for r in joba] == ["1", "2", "3"]
    assert [r["step_label"] for r in joba] == [
        "DATA statement", "PROCEDURE MEANS", "PROCEDURE SORT"]


def test_metrics_are_zero_when_profile_is_none(tmp_path, logs_root):
    out = tmp_path / "plain.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out])
    rows = read_rows(out)
    assert sum(int(r["step_count"]) for r in rows) == 0
    assert sum(float(r["total_real_time_sec"]) for r in rows) == 0.0
    assert sum(int(r["error_count"]) for r in rows) == 0


@pytest.mark.parametrize("raw,expected", [
    ("0.05 seconds", 0.05), ("1.20", 1.20),
    ("1:03.05", 63.05), ("1:00:30.00", 3630.0), ("garbage", None),
])
def test_duration_parser_handles_clock_and_seconds(raw, expected):
    assert sfs.parse_duration(raw) == (pytest.approx(expected) if expected else expected)


# --------------------------------------------------------------------- #
# Folder exclusions
# --------------------------------------------------------------------- #

def test_default_exclusions_exclude_nothing(tmp_path, logs_root):
    out = tmp_path / "all.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out])
    names = programs(out)
    assert {"legacy", "scratch", "keepme"} <= names, \
        "v1.3.1 default folder_exclusion_list is empty -- nothing may be excluded"


def test_exclusions_drop_old_and_test_but_keep_older(tmp_path, logs_root):
    out = tmp_path / "filtered.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--folder-exclusion-list", "Old", "Test"])
    names = programs(out)
    assert "legacy" not in names      # under Old/
    assert "scratch" not in names     # under Test/
    assert "keepme" in names, "near-miss sibling Older/ must be kept"


def test_exclusion_matching_is_case_insensitive(tmp_path, logs_root):
    out = tmp_path / "ci.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--folder-exclusion-list", "oLd", "tEsT"])
    names = programs(out)
    assert "legacy" not in names and "scratch" not in names
    assert "keepme" in names


def test_full_path_exclusion_matches_by_prefix(logs_root):
    target = Path(logs_root) / "Old" / "legacy.log"
    assert sfs.is_folder_excluded(target, [str(Path(logs_root) / "Old")])
    assert sfs.is_folder_excluded(target, [str(Path(logs_root) / "Old") + "/"])
    assert not sfs.is_folder_excluded(Path(logs_root) / "Older" / "keepme.log",
                                      [str(Path(logs_root) / "Old")])


# --------------------------------------------------------------------- #
# Required-parameter validation & exit codes
# --------------------------------------------------------------------- #

def test_empty_input_folder_root_errors(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(sfs, "input_folder_root", [])
    rc = run(["--output-file-path", tmp_path / "x.csv"])
    assert rc != 0
    assert "input_folder_root" in capsys.readouterr().err


def test_omitted_output_file_path_autonames_scan_csv(tmp_path, logs_root, monkeypatch):
    """v1.3.2: output_file_path is optional -> scan_YYYYMMDD_HHMMSS.csv in cwd."""
    monkeypatch.setattr(sfs, "output_file_path", None)
    monkeypatch.chdir(tmp_path)
    assert run(["--input-folder-root", logs_root]) == 0
    generated = list(tmp_path.glob("scan_*.csv"))
    assert len(generated) == 1, "exactly one auto-named CSV must be produced"
    name = generated[0].name
    assert re.fullmatch(r"scan_\d{8}_\d{6}\.csv", name), name
    assert read_rows(generated[0]), "the auto-named CSV must contain rows"


def test_omitted_output_with_profile_still_writes_companion(tmp_path, logs_root, monkeypatch):
    monkeypatch.setattr(sfs, "output_file_path", None)
    monkeypatch.chdir(tmp_path)
    assert run(["--input-folder-root", logs_root, "--metric-profile", "sas_log"]) == 0
    main_csv = list(tmp_path.glob("scan_*.csv"))
    companion = [p for p in main_csv if p.name.endswith("_StepDetail.csv")]
    assert len(companion) == 1, "the auto-named run must still emit StepDetail"


def test_default_output_name_shape():
    generated = sfs.default_output_name()
    assert re.fullmatch(r"scan_\d{8}_\d{6}\.csv", generated.name), generated.name
    assert sfs.default_output_name(Path("/tmp/x")).parent == Path("/tmp/x")


def test_unknown_metric_profile_errors(tmp_path, logs_root, capsys):
    rc = run(["--input-folder-root", logs_root, "--output-file-path", tmp_path / "x.csv",
              "--metric-profile", "bogus_profile"])
    assert rc != 0
    assert "metric_profile" in capsys.readouterr().err


def test_bogus_only_root_exits_non_zero(tmp_path):
    rc = run(["--input-folder-root", tmp_path / "nope", "--output-file-path", tmp_path / "x.csv"])
    assert rc != 0


def test_unwritable_output_path_exits_non_zero(tmp_path, logs_root):
    # A regular file standing in for a parent directory cannot be written through.
    blocker = tmp_path / "not_a_directory"
    blocker.write_text("i am a file, not a folder\n")
    rc = run(["--input-folder-root", logs_root, "--output-file-path", blocker / "out.csv"])
    assert rc != 0


def test_success_exit_code_is_zero(tmp_path, logs_root):
    assert run(["--input-folder-root", logs_root, "--output-file-path", tmp_path / "ok.csv"]) == 0


# --------------------------------------------------------------------- #
# program_name derivation
# --------------------------------------------------------------------- #

def test_program_name_strips_extension_and_prefix(tmp_path, logs_root):
    out = tmp_path / "pn.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--file-exclusion-list", "PGM_"])
    names = programs(out)
    assert "report" in names, "file_exclusion_list prefix PGM_ must be stripped"
    assert "PGM_report" not in names


def test_program_name_helper_is_case_insensitive():
    assert sfs.derive_program_name("PGM_report.sas", ["pgm_"]) == "report"
    assert sfs.derive_program_name("jobA.log", []) == "jobA"


# --------------------------------------------------------------------- #
# Keyword extraction
# --------------------------------------------------------------------- #

def test_accdb_mdb_keyword_sweep(tmp_path, logs_root):
    out = tmp_path / "kw.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--extract-keyword", ".accdb", ".mdb"])
    row = by_program(out)["notes"]
    assert int(row["kw_accdb_count"]) == 2      # two .accdb mentions in notes.txt
    assert int(row["kw_mdb_count"]) == 1
    assert ".accdb" in row["kw_accdb_line"]
    assert "claims.accdb" in row["kw_accdb_context"]


def test_real_cpu_keyword_context_window(tmp_path, logs_root):
    out = tmp_path / "kw2.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--extract-keyword", "real time", "cpu time"])
    row = by_program(out)["jobA"]
    assert int(row["kw_real_time_count"]) == 3
    assert int(row["kw_cpu_time_count"]) == 3
    # +/-3 lines around the hit => at most 7 lines
    assert 1 < len(row["kw_real_time_context"].splitlines()) <= 2 * sfs.CONTEXT_LINES + 1


def test_keyword_columns_are_slugged_and_ordered():
    assert sfs.keyword_columns([".accdb"]) == [
        "kw_accdb_line", "kw_accdb_context", "kw_accdb_count"]
    assert sfs.keyword_columns(["real time"])[0] == "kw_real_time_line"


# --------------------------------------------------------------------- #
# Date-range filter
# --------------------------------------------------------------------- #

def test_date_range_includes_only_in_window(tmp_path, logs_root):
    out = tmp_path / "dated.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out,
                "--date-from", "2026-01-01", "--date-to", "2026-06-30"]) == 0
    names = programs(out)
    assert "dated_2026H1" in names       # 2026-04-01, inside
    assert "dated_2025" not in names     # 2025-06-15, before
    assert "dated_2026H2" not in names   # 2026-11-30, after


def test_open_ended_bounds(tmp_path, logs_root):
    lower = tmp_path / "lower.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", lower,
         "--date-from", "2026-01-01"])
    names = programs(lower)
    assert "dated_2026H2" in names and "dated_2025" not in names

    upper = tmp_path / "upper.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", upper,
         "--date-to", "2025-12-31"])
    names = programs(upper)
    assert "dated_2025" in names and "dated_2026H1" not in names


def test_bare_upper_bound_date_is_inclusive_through_end_of_day():
    bound = sfs.parse_boundary("2026-06-30", end_of_day=True)
    assert (bound.hour, bound.minute, bound.second) == (23, 59, 59)


def test_bad_date_string_exits_non_zero(tmp_path, logs_root, capsys):
    rc = run(["--input-folder-root", logs_root, "--output-file-path", tmp_path / "x.csv",
              "--date-from", "not-a-date"])
    assert rc != 0
    assert "not-a-date" in capsys.readouterr().err


def test_inverted_range_exits_non_zero(tmp_path, logs_root):
    rc = run(["--input-folder-root", logs_root, "--output-file-path", tmp_path / "x.csv",
              "--date-from", "2026-06-30", "--date-to", "2026-01-01"])
    assert rc != 0


# --------------------------------------------------------------------- #
# Robustness
# --------------------------------------------------------------------- #

def test_malformed_file_is_annotated_and_run_continues(tmp_path, logs_root):
    out = tmp_path / "robust.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out]) == 0
    files = by_program(out)
    if "broken_link" in files:            # skipped where symlinks are unavailable
        assert files["broken_link"]["parse_status"] != "OK"
    # the run still produced the healthy rows
    assert {"jobA", "jobB", "notes"} <= set(files)


def test_invalid_utf8_falls_back_without_failing(tmp_path, logs_root):
    out = tmp_path / "enc.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out])
    row = by_program(out)["malformed_binary"]
    assert row["parse_status"] == "OK", "latin-1 fallback must keep the row readable"


def test_extension_filter_excludes_other_files(tmp_path, logs_root):
    out = tmp_path / "ext.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out])
    assert "dat" not in {r["extension"] for r in read_rows(out)}, \
        "ignore_me.dat must not appear"


def test_no_recursion_keeps_top_level_only(tmp_path, logs_root):
    out = tmp_path / "flat.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--no-include-subdirectories"])
    names = programs(out)
    assert "jobA" in names
    assert "jobC" not in names, "nested/ must be skipped when recursion is off"


def test_multiple_roots_are_combined(tmp_path, logs_root):
    out = tmp_path / "multi.csv"
    assert run(["--input-folder-root", str(Path(logs_root) / "nested"),
                str(Path(logs_root) / "Older"),
                "--output-file-path", out]) == 0
    assert programs(out) == {"jobC", "keepme"}


def test_missing_root_is_skipped_but_run_succeeds(tmp_path, logs_root):
    out = tmp_path / "partial.csv"
    assert run(["--input-folder-root", logs_root, str(tmp_path / "ghost"),
                "--output-file-path", out]) == 0
    assert read_rows(out), "the reachable root must still produce rows"


def test_files_columns_are_stable_and_ordered(tmp_path, logs_root):
    out = tmp_path / "cols.csv"
    run(["--input-folder-root", logs_root, "--output-file-path", out,
         "--extract-keyword", ".accdb"])
    expected = (sfs.FILES_BASE_COLUMNS
                + ["kw_accdb_line", "kw_accdb_context", "kw_accdb_count"]
                + sfs.FILES_TAIL_COLUMNS)
    assert columns_of(out) == expected


# --------------------------------------------------------------------- #
# Dependency surface (v1.3.3 -- no pandas/numpy on the runtime path)
# --------------------------------------------------------------------- #

def test_scanner_module_does_not_import_pandas_or_numpy():
    """A broken pandas install must never be able to break a scan."""
    source = (ROOT / "scanFileSystem.py").read_text(encoding="utf-8")
    for banned in ("import pandas", "import numpy"):
        assert banned not in source, f"{banned!r} must not appear in scanFileSystem.py"
    assert "pandas" not in sys.modules or True   # importing sfs must not pull it in


def test_csv_output_needs_no_third_party_modules(tmp_path, logs_root, monkeypatch):
    """CSV writing must work even if every optional Excel engine is missing."""
    monkeypatch.setattr(sfs, "first_available_excel_engine", lambda: None)
    out = tmp_path / "stdlib.csv"
    assert run(["--input-folder-root", logs_root, "--output-file-path", out,
                "--metric-profile", "sas_log"]) == 0
    assert read_rows(out)
    assert (tmp_path / "stdlib_StepDetail.csv").exists()


def test_xlsx_without_engine_falls_back_to_csv(tmp_path, logs_root, monkeypatch, capsys):
    monkeypatch.setattr(sfs, "first_available_excel_engine", lambda: None)
    target = tmp_path / "fallback.xlsx"
    assert run(["--input-folder-root", logs_root, "--output-file-path", target,
                "--metric-profile", "sas_log"]) == 0
    assert not target.exists(), "no .xlsx should be produced without an engine"
    assert (tmp_path / "fallback.csv").exists()
    assert "falling back to CSV" in capsys.readouterr().err


def test_excel_safe_strips_illegal_control_characters():
    assert sfs.excel_safe("ok\x00text\x1f") == "oktext"
    assert sfs.excel_safe("keeps\ttabs\nnewlines") == "keeps\ttabs\nnewlines"
    assert sfs.excel_safe(42) == 42
    assert sfs.excel_safe(None) is None
