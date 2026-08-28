"""Data-quality and validation tests across the cgs_ai functions."""
from __future__ import annotations
import sys
from pathlib import Path
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from src.py.copyExcelSheet2CSV import SheetNotCsvReadyError, copyExcelSheet2CSV  # noqa: E402
from src.py.downloadBulkFiles import downloadBulkFiles  # noqa: E402
from src.py.formatCSV import formatCSV                  # noqa: E402
from src.py.scanFileSystem import scanFileSystem        # noqa: E402
from src.py.sendEmail import sendEmail                  # noqa: E402
from src.py.zipFolder import zipFolder                  # noqa: E402
from src.utils.config import parseEnvFile               # noqa: E402

openpyxl = pytest.importorskip("openpyxl", reason="Excel tests need openpyxl")


# --- required-parameter validation -------------------------------------------

@pytest.mark.parametrize("call,missing", [
    (lambda: sendEmail(To="", From="a@b.c", Subject="s", Body="b"), "To"),
    (lambda: sendEmail(To="a@b.c", From="", Subject="s", Body="b"), "From"),
    (lambda: zipFolder(FolderToZip="", OutputZipPath="x.zip"), "FolderToZip"),
    (lambda: zipFolder(FolderToZip=".", OutputZipPath=""), "OutputZipPath"),
    (lambda: downloadBulkFiles(InputCsvPath="", OutputFolder="o"), "InputCsvPath"),
    (lambda: formatCSV(InputCsvPath="", OutputExcelPath="o.xlsx"), "InputCsvPath"),
])
def test_required_parameters_are_enforced(call, missing):
    with pytest.raises(ValueError) as excinfo:
        call()
    assert missing in str(excinfo.value)


def test_unknown_metric_profile_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="metric_profile"):
        scanFileSystem(input_folder_root=str(tmp_path), extract_keyword=["x"],
                       metric_profile="bogus")


def test_bad_date_is_rejected(tmp_path):
    with pytest.raises(ValueError, match="unparseable date"):
        scanFileSystem(input_folder_root=str(tmp_path), extract_keyword=["x"],
                       date_from="not-a-date")


# --- Excel validation --------------------------------------------------------

def _workbook(tmp_path, rows, name="wb.xlsx", sheet="Data"):
    from openpyxl import Workbook
    workbook = Workbook()
    sheetObj = workbook.active
    sheetObj.title = sheet
    for row in rows:
        sheetObj.append(row)
    path = tmp_path / name
    workbook.save(path)
    return str(path)


def test_missing_sheet_is_refused(tmp_path):
    path = _workbook(tmp_path, [["a", "b"], [1, 2]])
    with pytest.raises(SheetNotCsvReadyError, match="not found"):
        copyExcelSheet2CSV(path, "NoSuchSheet", str(tmp_path / "o.csv"))


def test_empty_sheet_is_refused(tmp_path):
    path = _workbook(tmp_path, [])
    with pytest.raises(SheetNotCsvReadyError, match="empty"):
        copyExcelSheet2CSV(path, "Data", str(tmp_path / "o.csv"))


def test_duplicate_headers_are_refused(tmp_path):
    path = _workbook(tmp_path, [["a", "a"], [1, 2]])
    with pytest.raises(SheetNotCsvReadyError, match="duplicate"):
        copyExcelSheet2CSV(path, "Data", str(tmp_path / "o.csv"))


def test_blank_header_is_refused(tmp_path):
    path = _workbook(tmp_path, [["a", None, "c"], [1, 2, 3]])
    with pytest.raises(SheetNotCsvReadyError, match="blank header"):
        copyExcelSheet2CSV(path, "Data", str(tmp_path / "o.csv"))


def test_nothing_is_written_when_validation_fails(tmp_path):
    path = _workbook(tmp_path, [["a", "a"], [1, 2]])
    out = tmp_path / "must_not_exist.csv"
    with pytest.raises(SheetNotCsvReadyError):
        copyExcelSheet2CSV(path, "Data", str(out))
    assert not out.exists(), "validation must refuse BEFORE writing"


# --- downloadBulkFiles data handling -----------------------------------------

def test_blank_link_cells_are_skipped_not_failed(tmp_path):
    csvPath = tmp_path / "links.csv"
    csvPath.write_text("commentId,attachmentLinks\nA,\nB,\nC,\n", encoding="utf-8")
    result = downloadBulkFiles(InputCsvPath=str(csvPath),
                               OutputFolder=str(tmp_path / "out"))
    assert result["Skipped"] == 3 and result["Failed"] == 0


def test_missing_link_column_names_the_available_columns(tmp_path):
    csvPath = tmp_path / "links.csv"
    csvPath.write_text("commentId,other\nA,x\n", encoding="utf-8")
    with pytest.raises(ValueError, match="Available columns"):
        downloadBulkFiles(InputCsvPath=str(csvPath),
                          OutputFolder=str(tmp_path / "out"))


# --- .env parsing ------------------------------------------------------------

def test_env_parsing_rules(tmp_path):
    envFile = tmp_path / ".env"
    envFile.write_text(
        "# a comment\n\nKEY=value\nSPACED = spaced\n"
        'QUOTED="quoted"\nWINPATH=\\\\srv\\share\\x\nNOEQUALS\n',
        encoding="utf-8")
    values = parseEnvFile(envFile)
    assert values["KEY"] == "value"
    assert values["SPACED"] == "spaced"        # tolerated even though discouraged
    assert values["QUOTED"] == "quoted"        # surrounding quotes stripped
    assert values["WINPATH"] == "\\\\srv\\share\\x"   # backslashes preserved
    assert "NOEQUALS" not in values
    assert not any(k.startswith("#") for k in values)


# --- formatCSV: the corporatev2 style ----------------------------------------

import runpy                                             # noqa: E402
from src.py.formatCSV import FORMAT_STYLES, STRIPED_FORMATS  # noqa: E402


def _sampleCsv(tmp_path):
    """Write a small CSV and return its path. Parameters: tmp_path (Path)."""
    path = tmp_path / "claims.csv"
    path.write_text("ClaimId,Amount\nA1,10\nA2,20\nA3,30\n", encoding="utf-8")
    return path


@pytest.mark.parametrize("style", ["corporate", "corporatev2", "plain", "minimal"])
def test_every_declared_format_type_is_accepted(tmp_path, style):
    result = formatCSV(InputCsvPath=str(_sampleCsv(tmp_path)),
                       OutputExcelPath=str(tmp_path / f"{style}.xlsx"),
                       FormatType=style)
    assert result["FormatType"] == style
    assert result["RowCount"] == 3


def test_corporatev2_is_registered_and_striped():
    assert "corporatev2" in FORMAT_STYLES
    assert "corporatev2" in STRIPED_FORMATS
    # v2 is the versioned name for the corporate look; the palettes must match.
    assert FORMAT_STYLES["corporatev2"] == FORMAT_STYLES["corporate"]


def test_corporatev2_renders_banner_header_and_stripes(tmp_path):
    output = tmp_path / "v2.xlsx"
    formatCSV(InputCsvPath=str(_sampleCsv(tmp_path)),
              OutputExcelPath=str(output), FormatType="corporatev2",
              SheetName="Claims", Title="CMS Reporting and Analysis")
    sheet = openpyxl.load_workbook(output)["Claims"]
    assert sheet["A1"].value == "CMS Reporting and Analysis"
    assert sheet["A1"].fill.fgColor.rgb.endswith("1F3864")   # navy banner
    assert sheet["A2"].fill.fgColor.rgb.endswith("2E75B6")   # blue header
    assert sheet["A3"].fill.fgColor.rgb.endswith("DCE6F1")   # zebra stripe
    assert sheet.freeze_panes == "A3"
    assert sheet.auto_filter.ref == "A2:B5"


def test_plain_format_is_not_striped(tmp_path):
    output = tmp_path / "plain.xlsx"
    formatCSV(InputCsvPath=str(_sampleCsv(tmp_path)),
              OutputExcelPath=str(output), FormatType="plain")
    sheet = openpyxl.load_workbook(output).active
    assert sheet["A3"].fill.fgColor.rgb in ("00000000", None)


def test_unknown_format_type_names_the_valid_ones(tmp_path):
    with pytest.raises(ValueError) as excinfo:
        formatCSV(InputCsvPath=str(_sampleCsv(tmp_path)),
                  OutputExcelPath=str(tmp_path / "x.xlsx"), FormatType="bogus")
    assert "corporatev2" in str(excinfo.value)


def test_lite_build_offers_the_same_format_types():
    """The lite formatCSV must stay in step with the full one."""
    namespace = runpy.run_path(str(ROOT / "src" / "py" / "lite" / "__init__.py"),
                               run_name="cgs_ai_lite_test")
    assert namespace["FORMAT_STYLES"] == FORMAT_STYLES
    assert namespace["STRIPED_FORMATS"] == STRIPED_FORMATS
