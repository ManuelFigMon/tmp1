"""Tests for the filescan pipeline, package exports and cross-language parity."""
from __future__ import annotations
import csv, re, subprocess, sys
from pathlib import Path
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT.parent))          # so `import cgs_ai` works
sys.path.insert(0, str(ROOT))

import cgs_ai                                            # noqa: E402
from src.pipelines.filescan_pipeline import runFilescanPipeline  # noqa: E402

SAMPLE = """NOTE: DATA statement used (Total process time):
      real time           0.05 seconds
      cpu time            0.03 seconds
NOTE: PROCEDURE MEANS used (Total process time):
      real time           1.20 seconds
      cpu time            0.90 seconds
"""


@pytest.fixture
def logsRoot(tmp_path):
    (tmp_path / "jobA.log").write_text(SAMPLE, encoding="utf-8")
    return str(tmp_path)


# --- package surface ---------------------------------------------------------

def test_version_is_declared():
    assert cgs_ai.__version__ == "1.0beta"


@pytest.mark.parametrize("name", [
    "scanFileSystem", "runSQLServerQuery", "formatCSV", "downloadBulkFiles",
    "sendEmail", "convertSAS2Pandas", "copyExcelSheet2CSV",
    "collectSystemMetrics", "zipFolder", "runFilescanPipeline",
    "basic_hello", "personalized_hello", "detailed_hello"])
def test_every_function_is_exported(name):
    assert hasattr(cgs_ai, name), f"cgs_ai.{name} is missing"
    assert name in cgs_ai.__all__


def test_greeting_functions():
    assert cgs_ai.basic_hello() == "Hello, World!"
    assert cgs_ai.personalized_hello("  Manuel  ") == "Hello, Manuel!"
    assert cgs_ai.personalized_hello("") == "Hello, World!"
    assert cgs_ai.detailed_hello("pirate")["style_used"] == "pirate"
    assert "Ahoy" in cgs_ai.detailed_hello("pirate")["message"]
    assert cgs_ai.detailed_hello("unknown")["message"] == \
        cgs_ai.detailed_hello("friendly")["message"]


def test_importing_cgs_ai_needs_no_third_party_package():
    """The package must import with the standard library alone."""
    code = ("import sys;"
            "blocked=('pandas','numpy','openpyxl','pyodbc','psutil');"
            "[sys.modules.__setitem__(m, None) for m in blocked];"
            f"sys.path.insert(0, r'{ROOT.parent}');"
            "import cgs_ai; print(cgs_ai.__version__)")
    result = subprocess.run([sys.executable, "-c", code],
                            capture_output=True, text=True)
    assert result.returncode == 0, result.stderr
    assert "1.0beta" in result.stdout


# --- pipeline ----------------------------------------------------------------

def test_pipeline_runs_scan_and_reports_steps(logsRoot, tmp_path):
    result = runFilescanPipeline(
        input_folder_root=logsRoot, extract_keyword=["real time"],
        output_file_path=str(tmp_path / "scan.csv"), metric_profile="none")
    assert "scanFileSystem" in result["Steps"]
    assert len(result["Scan"]["matches"]) == 2


def test_pipeline_skips_email_when_no_recipient(logsRoot, tmp_path):
    result = runFilescanPipeline(
        input_folder_root=logsRoot, extract_keyword=["real time"],
        output_file_path=str(tmp_path / "scan.csv"), metric_profile="none")
    assert result["Email"] is None
    assert "sendEmail" not in result["Steps"]


def test_metric_profile_switches_output_to_excel(logsRoot, tmp_path):
    """A .csv request becomes .xlsx, because metrics need a second sheet."""
    pytest.importorskip("openpyxl")
    result = cgs_ai.scanFileSystem(
        input_folder_root=logsRoot, extract_keyword=["real time"],
        output_file_path=str(tmp_path / "scan.csv"), metric_profile="sas_log")
    assert result["output"].endswith(".xlsx")
    assert len(result["metrics"]) == 2
    from openpyxl import load_workbook
    assert load_workbook(result["output"]).sheetnames == ["Matches", "Metrics"]


def test_fullstimer_cpu_time_is_captured(tmp_path):
    """FULLSTIMER logs say 'user cpu time', not 'cpu time'."""
    (tmp_path / "fs.log").write_text(
        "NOTE: PROCEDURE SORT used (Total process time):\n"
        "      real time           12.34 seconds\n"
        "      user cpu time       0.80 seconds\n"
        "      system cpu time     0.20 seconds\n", encoding="utf-8")
    result = cgs_ai.scanFileSystem(
        input_folder_root=str(tmp_path), extract_keyword=["real time"],
        output_file_path=str(tmp_path / "o.csv"), metric_profile="none")
    from src.py.scanFileSystem import METRIC_PROFILES, parseMetricProfile
    lines = (tmp_path / "fs.log").read_text().splitlines()
    rows = parseMetricProfile(METRIC_PROFILES["sas_log"], lines, "p", "fs")
    assert rows[0]["CpuTimeSec"] == pytest.approx(0.80), \
        "FULLSTIMER 'user cpu time' must be captured, not silently dropped"


# --- the access_db profile ---------------------------------------------------

def accessDbRows(lines):
    """Run the access_db extractor over a list of lines. Returns: list[dict]."""
    from src.py.scanFileSystem import METRIC_PROFILES, parseMetricProfile
    return parseMetricProfile(METRIC_PROFILES["access_db"], lines, "p.sas", "p")


def buildProgram(placements, total=60):
    """Lay text at 1-based line numbers so a test can assert exact numbers.

    Parameters: placements (dict) - {lineNumber: text}; total (int) - length.
    Returns: list[str] of `total` lines, unplaced ones being a filler comment.
    """
    lines = ["/* filler */"] * total
    for number, text in placements.items():
        lines[number - 1] = text
    return lines


def test_access_db_reports_every_line_the_libref_is_used_on():
    rows = accessDbRows(buildProgram({
        3: r'libname issuelog access path="\\srv\dbs\issuelog.mdb";',
        12: "  set issuelog.claims;",
        48: "  create table b as select * from issuelog.detail",
        49: "  where x in (select y from issuelog.lookup);",
    }))
    assert len(rows) == 1
    assert rows[0]["Libref"] == "issuelog"
    assert rows[0]["Keyword"] == "mdb"
    assert rows[0]["DefinitionLine"] == 3
    assert rows[0]["UsageCount"] == 3
    assert rows[0]["UsageLines"] == "12,48,49"


def test_access_db_does_not_count_the_definition_line_itself():
    """The .mdb file is usually named after the libref, so 'issuelog.mdb' on
    the LIBNAME line matches the libref-dot pattern. It is not a use."""
    rows = accessDbRows(buildProgram({
        3: r'libname issuelog access path="\\srv\dbs\issuelog.mdb";',
    }))
    assert rows[0]["UsageCount"] == 0, \
        "the libref appears in its own file name, which is not a usage"


def test_access_db_finds_a_libname_wrapped_across_lines():
    rows = accessDbRows(buildProgram({
        7: "libname fmrrpt access",
        8: r'    path="\\srv\dbs\fmrrpt RHHI.mdb";',
        20: "  set fmrrpt.detail;",
    }))
    assert len(rows) == 1
    assert rows[0]["DefinitionLine"] == 7, "the line the statement STARTS on"
    assert rows[0]["DatabaseFile"].endswith("fmrrpt RHHI.mdb")
    assert rows[0]["UsageLines"] == "20"


def test_access_db_reports_na_when_defined_but_never_used():
    rows = accessDbRows(buildProgram({
        3: r'libname unused access path="c:\tmp\never.accdb";',
    }))
    assert rows[0]["UsageCount"] == 0
    assert rows[0]["UsageLines"] == "NA"
    assert rows[0]["Keyword"] == "accdb"


def test_access_db_ignores_a_macro_variable_of_the_same_name():
    """`&fmrrpt` and `%let fmrrpt` are not uses of the library."""
    rows = accessDbRows(buildProgram({
        3: r'libname fmrrpt access path="\\srv\dbs\fmrrpt.mdb";',
        10: "%let fmrrpt = 1;",
        11: "  if x = &fmrrpt then delete;",
        12: "  set fmrrpt.detail;",
    }))
    assert rows[0]["UsageLines"] == "12", \
        "only libref-DOT counts; a bare word collides with a macro variable"


def test_access_db_does_not_count_a_libname_clear_statement():
    rows = accessDbRows(buildProgram({
        3: r'libname issuelog access path="\\srv\dbs\issuelog.mdb";',
        12: "  set issuelog.claims;",
        40: "libname issuelog clear;",
    }))
    assert rows[0]["UsageLines"] == "12"


def test_access_db_reports_mdb_and_accdb_in_the_same_file():
    rows = accessDbRows(buildProgram({
        3: r'libname older access path="\\srv\dbs\legacy.mdb";',
        4: r'libname newer access path="\\srv\dbs\current.accdb";',
        20: "  set older.claims;",
        21: "  set newer.claims;",
    }))
    assert [(r["Libref"], r["Keyword"], r["UsageLines"]) for r in rows] == [
        ("older", "mdb", "20"), ("newer", "accdb", "21")]


def test_access_db_skips_a_macro_built_path_instead_of_misreporting_it():
    """'&mdbPath' contains the literal 'mdb'. A substring match would report
    this as a genuine Access library; it must be counted as skipped."""
    rows = accessDbRows(buildProgram({
        3: 'libname db pcfiles path="&mdbPath";',
        12: "  set db.claims;",
    }))
    assert rows == [], "a macro-built path has no visible .mdb/.accdb"


def test_access_db_does_not_warn_about_an_ordinary_sas_library():
    """A plain library with a macro in its path is not an Access candidate."""
    from src.py import scanFileSystem as scanner
    warnings = []
    original = scanner.logWarn
    scanner.logWarn = warnings.append
    try:
        accessDbRows(buildProgram({3: r'libname plain "&root\sasdata";'}))
    finally:
        scanner.logWarn = original
    assert warnings == [], "no access/pcfiles engine, so nothing was skipped"


def test_access_db_switches_output_to_excel_with_its_own_columns(tmp_path):
    pytest.importorskip("openpyxl")
    from src.py.scanFileSystem import ACCESS_DB_COLUMNS
    (tmp_path / "job.sas").write_text(
        'libname issuelog access path="\\\\srv\\dbs\\issuelog.mdb";\n'
        "data a; set issuelog.claims; run;\n", encoding="utf-8")
    result = cgs_ai.scanFileSystem(
        input_folder_root=str(tmp_path), extract_keyword=["mdb"],
        file_extensions=["sas"], output_file_path=str(tmp_path / "o.csv"),
        metric_profile="access_db")
    assert result["output"].endswith(".xlsx")
    assert len(result["metrics"]) == 1
    from openpyxl import load_workbook
    book = load_workbook(result["output"])
    assert book.sheetnames == ["Matches", "Metrics"]
    assert [c.value for c in book["Metrics"][1]] == ACCESS_DB_COLUMNS


def test_the_sas_log_profile_is_unchanged_by_the_registry_refactor():
    from src.py.scanFileSystem import (METRIC_COLUMNS, METRIC_PROFILES,
                                       SAS_LOG_COLUMNS, parseMetricProfile)
    assert METRIC_COLUMNS == SAS_LOG_COLUMNS, "the old name must still work"
    assert METRIC_PROFILES["sas_log"]["Columns"] == [
        "FullPath", "ProgramName", "StepIndex", "StepLabel",
        "RealTimeSec", "CpuTimeSec"]
    rows = parseMetricProfile(METRIC_PROFILES["sas_log"],
                              SAMPLE.splitlines(), "p", "job")
    assert [(r["StepIndex"], r["StepLabel"], r["RealTimeSec"]) for r in rows] == [
        (1, "DATA statement", 0.05), (2, "PROCEDURE MEANS", 1.20)]


# --- cross-language parity ---------------------------------------------------

PS_DIR = ROOT / "src" / "ps"
PY_DIR = ROOT / "src" / "py"
FUNCTIONS = ["scanFileSystem", "runSQLServerQuery", "formatCSV",
             "downloadBulkFiles", "sendEmail", "convertSAS2Pandas",
             "copyExcelSheet2CSV", "collectSystemMetrics", "zipFolder"]


@pytest.mark.parametrize("name", FUNCTIONS)
def test_every_function_exists_in_both_languages(name):
    assert (PY_DIR / f"{name}.py").is_file(), f"missing src/py/{name}.py"
    assert (PS_DIR / f"{name}.ps1").is_file(), f"missing src/ps/{name}.ps1"


@pytest.mark.parametrize("name", FUNCTIONS)
def test_every_function_has_the_standard_header(name):
    """Every file carries the scanFileSystem header block."""
    for path in (PY_DIR / f"{name}.py", PS_DIR / f"{name}.ps1"):
        text = path.read_text(encoding="utf-8")[:3000]
        for field in ("Program Name", "Author", "Purpose", "Version",
                      "Dependencies", "Input Parameters"):
            assert field in text, f"{path.name} header is missing '{field}'"


@pytest.mark.parametrize("name", FUNCTIONS)
def test_sas_wrapper_exists_for_every_function(name):
    sasText = (ROOT / "src" / "sas" / "cgsFunctions.sas").read_text()
    assert re.search(rf"%macro\s+{name}\s*\(", sasText), \
        f"cgsFunctions.sas has no %{name} wrapper"


def test_metric_profiles_match_across_languages():
    """The two engines must offer the same profiles with the same columns.

    PowerShell cannot be executed here, so this compares the DECLARATIONS --
    profile names, column lists and the parsing patterns -- which is what
    actually drifts when a profile is added to one engine only.
    """
    from src.py.scanFileSystem import (ACCESS_DB_COLUMNS, METRIC_PROFILES,
                                       SAS_LOG_COLUMNS)
    psText = (PS_DIR / "scanFileSystem.ps1").read_text()

    registry = re.search(r"\$script:MetricProfiles = @\{(.*?)\n\}\n",
                         psText, re.DOTALL).group(1)
    psNames = set(re.findall(r"^\s*'(\w+)'\s*=\s*@\{", registry, re.MULTILINE))
    assert psNames == set(METRIC_PROFILES), \
        f"PowerShell profiles {psNames} != Python {set(METRIC_PROFILES)}"

    for name, columns in (("SasLogColumns", SAS_LOG_COLUMNS),
                          ("AccessDbColumns", ACCESS_DB_COLUMNS)):
        block = re.search(rf"\$script:{name}\s*=\s*@\((.*?)\)\n",
                          psText, re.DOTALL).group(1)
        assert re.findall(r"'([^']+)'", block) == columns, \
            f"$script:{name} does not match its Python twin"

    # The patterns that decide what counts as a hit must be identical too.
    for pattern in (r"\blibname\s+(?<libref>[A-Za-z_]\w*)",
                    r"\b(?:access|pcfiles)\b"):
        assert pattern in psText, f"PowerShell is missing the pattern {pattern}"


def test_scanfilesystem_parameter_names_match_across_languages():
    """Parameter names must be identical in Python, PowerShell and SAS."""
    expected = {
        "input_folder_root", "extract_keyword", "output_file_path",
        "file_extensions", "include_subdirectories", "folder_exclusion_list",
        "file_exclusion_list", "lines_above", "lines_below",
        "nth_token_after", "nth_token_before", "numeric_token_after",
        "date_from", "date_to", "date_field", "metric_profile"}

    import inspect
    pySig = set(inspect.signature(cgs_ai.scanFileSystem).parameters) - {"self"}
    assert expected <= pySig, f"Python is missing {expected - pySig}"

    psText = (PS_DIR / "scanFileSystem.ps1").read_text()
    psParams = set(re.findall(r"\$(\w+)\s*=", psText.split("Set-StrictMode")[0]))
    assert expected <= psParams, f"PowerShell is missing {expected - psParams}"

    sasText = (ROOT / "src" / "sas" / "cgsFunctions.sas").read_text()
    block = re.search(r"%macro\s+scanFileSystem\s*\((.*?)\);", sasText, re.DOTALL).group(1)
    block = re.sub(r"/\*.*?\*/", "", block, flags=re.DOTALL)
    sasParams = set(re.findall(r"(\w+)\s*=", block))
    assert expected <= sasParams, f"SAS is missing {expected - sasParams}"


# --- formatCSV: the SAS ODS1 renderer and the PowerShell native writer -------

SAS_FUNCTIONS = (ROOT / "src" / "sas" / "cgsFunctions.sas").read_text()
PS_FORMATCSV = (ROOT / "src" / "ps" / "formatCSV.ps1").read_text()


def test_sas_macro_and_mend_are_balanced():
    """Every %macro needs its %mend, or everything after it is swallowed."""
    opens = re.findall(r"^%macro\s+(\w+)", SAS_FUNCTIONS, re.MULTILINE)
    closes = re.findall(r"^%mend\s+(\w+)", SAS_FUNCTIONS, re.MULTILINE)
    assert opens == closes, f"unbalanced: {set(opens) ^ set(closes)}"


def test_sas_formatcsv_routes_ods1_without_leaving_sas():
    block = re.search(r"%macro formatCSV\b.*?%mend formatCSV;",
                      SAS_FUNCTIONS, re.DOTALL).group(0)
    # ODS1 must be intercepted BEFORE %cgsRun, or it would be handed to
    # PowerShell, which has no such FormatType.
    assert "%upcase(&FormatType) = ODS1" in block
    assert block.index("ODS1") < block.index("%cgsRun")
    assert "%cgsFormatCsvOds" in block
    # The parameter comment must list ODS1, or callers never learn it exists.
    header = block[:block.index(");")]
    assert "ODS1" in header, "the FormatType comment does not mention ODS1"


def test_sas_ods1_renderer_is_self_contained():
    block = re.search(r"%macro cgsFormatCsvOds\b.*?%mend cgsFormatCsvOds;",
                      SAS_FUNCTIONS, re.DOTALL).group(0)
    for needed in ("ods excel", "proc report", "ods excel close",
                   "cx1F3864", "cx2E75B6", "cxDCE6F1", "frozen_headers",
                   "autofilter"):
        assert needed in block, f"ODS1 renderer is missing {needed!r}"
    # It must not shell out; that is the whole point of ODS1.
    assert "%cgsRun" not in block
    # PROC IMPORT samples the file to guess types and fails outright on a
    # header-only CSV -- which is what scanFileSystem writes when nothing
    # matches. The DATA step reader does not sample. Check the CODE only:
    # the comments in this macro discuss PROC IMPORT by name.
    code = re.sub(r"/\*.*?\*/", "", block, flags=re.DOTALL).lower()
    assert "proc import" not in code, \
        "ODS1 must not use PROC IMPORT; it cannot read a header-only CSV"
    assert "infile" in code and "firstobs=2" in code


def test_sas_ods1_handles_a_header_only_csv():
    """scanFileSystem writes a header and no rows when nothing matches."""
    block = re.search(r"%macro cgsFormatCsvOds\b.*?%mend cgsFormatCsvOds;",
                      SAS_FUNCTIONS, re.DOTALL).group(0)
    # It must count the rows and say so rather than failing.
    assert "nlobs" in block
    assert "no data rows" in block
    # A truly empty file is a different, clearly reported condition.
    assert "no header row" in block


def test_sas_ods1_reads_every_column_as_character():
    """Matches the Python and PowerShell twins, and keeps leading zeros."""
    block = re.search(r"%macro cgsFormatCsvOds\b.*?%mend cgsFormatCsvOds;",
                      SAS_FUNCTIONS, re.DOTALL).group(0)
    assert "length &_cgsVarList $ &ColumnLength;" in block
    # A one-column CSV cannot be written as the range _c1-_c1.
    assert "%if &_cgsN = 1 %then %let _cgsVarList = _c1;" in block
    # open() must be paired with close() or the dataset id leaks.
    assert block.count("%sysfunc(open(") == block.count("%sysfunc(close(")


#: Tokens a SAS/macro statement may legitimately begin with.
SAS_STATEMENT_STARTS = (
    "%", "/*", "*", ";", "data ", "proc ", "run;", "quit;", "ods ", "title",
    "footnote", "libname", "filename", "options", "label", "column", "define",
    "compute", "endcomp", "if ", "call ", "set ", "infile", "input", "length",
    "end;", "do ", "select", "attrib", "format", "keep", "drop", "where",
)


def test_sas_put_statements_have_no_embedded_semicolons():
    """A ';' inside %put text ends the statement early -- ERROR 180-322."""
    for line in SAS_FUNCTIONS.splitlines():
        stripped = line.strip()
        if stripped.startswith("%put"):
            body = stripped[len("%put"):].rstrip()
            assert body.count(";") == 1 and body.endswith(";"), \
                f"%put with an embedded semicolon: {stripped}"


def test_sas_put_statements_are_not_continued_onto_the_next_line():
    """A %put that ends with ';' cannot be continued.

    Text on the following line is no longer part of the message; it becomes
    stray macro text that SAS tries to execute. This is the same defect as
    an embedded semicolon, one line further down, and the single-line check
    above does not see it.
    """
    lines = SAS_FUNCTIONS.splitlines()
    for index, line in enumerate(lines):
        if not line.strip().startswith("%put"):
            continue
        if not line.strip().endswith(";"):
            continue
        for following in lines[index + 1:]:
            nextLine = following.strip()
            if not nextLine:
                break
            assert nextLine.lower().startswith(SAS_STATEMENT_STARTS), (
                f"line {index + 2} continues a finished %put: {nextLine!r}")
            break


def test_powershell_formatcsv_no_longer_fails_without_importexcel():
    """The missing module must be a fallback, not a fatal error."""
    assert "formatCSV requires the ImportExcel module" not in PS_FORMATCSV
    assert "function Write-XlsxNative" in PS_FORMATCSV
    assert "writing the workbook natively" in PS_FORMATCSV


def test_powershell_native_writer_emits_valid_ooxml_shape():
    # autoFilter must precede mergeCells; Excel refuses the other order.
    assert PS_FORMATCSV.index("'<autoFilter ref=") < PS_FORMATCSV.index("'<mergeCells count=")
    # Fill 0 must be none and fill 1 gray125, or Excel rejects the workbook.
    fills = re.search(r'<fills count="5">(.*?)</fills>', PS_FORMATCSV, re.DOTALL).group(1)
    assert fills.index('patternType="none"') < fills.index('patternType="gray125"')
    # A BOM inside a part makes the file unreadable.
    assert "UTF8Encoding($false)" in PS_FORMATCSV
    assert "<cellStyles count=" in PS_FORMATCSV


def test_powershell_writer_switch_is_documented_and_validated():
    assert "[string] $Writer" in PS_FORMATCSV
    assert "'auto', 'native', 'module'" in PS_FORMATCSV


# --- SAS block comments must not nest ----------------------------------------

def _scanBlockComments(text):
    """Return (nested_line_numbers, unterminated).

    SAS block comments do NOT nest: the first close marker ends the comment,
    so a nested open means everything after that close is parsed as live code.
    """
    depth, index, line, nested = 0, 0, 1, []
    while index < len(text) - 1:
        pair = text[index:index + 2]
        if pair == "/*":
            if depth:
                nested.append(line)
            else:
                depth = 1
            index += 2
            continue
        if pair == "*/" and depth:
            depth = 0
            index += 2
            continue
        if text[index] == "\n":
            line += 1
        index += 1
    return nested, bool(depth)


@pytest.mark.parametrize(
    "sasPath",
    sorted((ROOT / "src" / "sas").glob("*.sas")),
    ids=lambda p: p.name)
def test_sas_block_comments_do_not_nest(sasPath):
    nested, unterminated = _scanBlockComments(sasPath.read_text())
    assert not nested, (
        f"{sasPath.name}: block comment opened again at line(s) {nested} while "
        f"one was already open -- the outer comment ends early and the rest "
        f"of the file is parsed as code")
    assert not unterminated, f"{sasPath.name}: a block comment is never closed"


# --- cgsCore argument building ------------------------------------------------

CGS_CORE = (ROOT / "src" / "sas" / "cgsCore.sas").read_text()


def test_cgscore_does_not_separate_arguments_with_a_semicolon():
    """';' is the delimiter INSIDE list values, so it cannot also separate
    arguments. When it did, extract_keyword=%str(real time;cpu time;accdb;mdb)
    became four command-line tokens and PowerShell bound the strays to
    whatever positional parameter came next -- surfacing as
    "unparseable date 'accdb'".
    """
    assert "scan(args, i, ';')" not in CGS_CORE
    assert "countw(args, ';')" not in CGS_CORE
    assert "%str(;)" not in CGS_CORE


def test_cgscore_uses_indexed_argument_macro_variables():
    assert "CGS_ARGNAME&CGS_ARG_N" in CGS_CORE
    assert "CGS_ARGVAL&CGS_ARG_N" in CGS_CORE
    assert "cats('CGS_ARGNAME', i)" in CGS_CORE
    assert "cats('CGS_ARGVAL', i)" in CGS_CORE


def test_cgscore_strips_line_breaks_from_argument_values():
    """A %str() spanning several lines otherwise splits the .bat in two."""
    assert "compress(symget(cats('CGS_ARGVAL', i)), '0D0A'x)" in CGS_CORE


def _buildCommand(arguments):
    """Mirror the cgsRun DATA step. Parameters: [(name, value, always)]."""
    recorded = [(name, value) for name, value, always in arguments
                if len(value) > 0 or always]
    command = ""
    for name, value in recorded:
        value = value.replace("\r", "").replace("\n", "").strip()
        command += f' {name} "{value}"' if value else f" {name}"
    return command.strip()


def test_semicolon_list_survives_as_one_command_line_argument():
    """The exact call from the field report."""
    command = _buildCommand([
        ("-input_folder_root", "\\\\srv\\HHH\\Old_logs;\n\t\\\\srv\\DME\\Logs", False),
        ("-extract_keyword", "real time;cpu time;accdb;mdb", False),
        ("-metric_profile", "sas_log", False),
    ])
    assert '-extract_keyword "real time;cpu time;accdb;mdb"' in command
    assert '-input_folder_root "\\\\srv\\HHH\\Old_logs;\t\\\\srv\\DME\\Logs"' in command
    # Every token is either a switch or lives inside quotes: no bare strays.
    outside = re.sub(r'"[^"]*"', "", command).split()
    assert all(token.startswith("-") for token in outside), \
        f"stray positional token(s) would bind to the wrong parameter: {outside}"


def test_blank_values_are_dropped_unless_always():
    command = _buildCommand([("-date_from", "", False),
                             ("-include_subdirectories", "1", True),
                             ("-flag", "", True)])
    assert "-date_from" not in command
    assert '-include_subdirectories "1"' in command
    assert command.endswith("-flag")


# --- the PowerShell twin checks the target too -------------------------------

PS_SCANNER = (ROOT / "src" / "ps" / "scanFileSystem.ps1").read_text()
PS_UTILS = (ROOT / "src" / "ps" / "cgsUtils.ps1").read_text()


def test_powershell_probes_the_target_before_crawling():
    assert "function Test-CgsWritable" in PS_UTILS
    assert "function Resolve-CgsWritableTarget" in PS_UTILS
    assert "$target = Resolve-CgsWritableTarget -Target $target" in PS_SCANNER
    # It must run before the roots are walked.
    assert PS_SCANNER.index("Resolve-CgsWritableTarget") < PS_SCANNER.index("foreach ($rawRoot in $roots)")


def test_powershell_probe_cleans_up_and_matches_the_writer():
    block = re.search(r"function Test-CgsWritable.*?\n}\n", PS_UTILS, re.DOTALL).group(0)
    # OpenOrCreate creates the file when absent; it must be removed again.
    assert "Remove-Item" in block and "$existed" in block
    # A stricter share mode would reject files the write would have accepted.
    assert "FileShare]::Read" in block
    assert "FileShare]::None" not in block


def test_powershell_falls_back_instead_of_losing_the_crawl():
    block = re.search(r"function Resolve-CgsWritableTarget.*?\n}\n",
                      PS_UTILS, re.DOTALL).group(0)
    assert "Get-CgsTimestampSuffix" in block, "the fallback needs a unique name"
    assert "open in Excel" in block
    # It warns and continues; it only throws when the fallback fails too.
    assert block.count("Write-CgsWarn") >= 2
    assert block.count("throw") == 1


def test_powershell_csv_writer_explains_a_lock():
    assert "cannot write '{0}'" in PS_UTILS
    assert PS_UTILS.count("open in Excel") >= 2


# --- every *-Cgs* helper a script calls must exist in src/ps -----------------

PS_FILES = sorted((ROOT / "src" / "ps").glob("*.ps1"))
#: Verb-Noun names in the shared cgsUtils namespace, e.g. Write-CgsWarn.
CGS_FUNCTION = re.compile(r"\b([A-Z][a-zA-Z]*-Cgs[A-Za-z]+)\b")


def _definedCgsFunctions():
    names = set()
    for path in PS_FILES:
        for match in re.finditer(r"^function\s+([A-Za-z]+-[A-Za-z]+)",
                                 path.read_text(), re.MULTILINE):
            names.add(match.group(1))
    return names


@pytest.mark.parametrize("psPath", PS_FILES, ids=lambda p: p.name)
def test_every_cgs_function_referenced_is_defined_in_src_ps(psPath):
    """Catch a rename that leaves a caller behind.

    The .ps1 files are dot-sourced from one folder and get copied to the
    share, so a helper renamed in cgsUtils.ps1 without its callers fails at
    RUNTIME with "the term X is not recognized" -- after the crawl, on the
    user's machine. This is that check, done at build time.
    """
    defined = _definedCgsFunctions()
    missing = sorted(set(CGS_FUNCTION.findall(psPath.read_text())) - defined)
    assert not missing, (
        f"{psPath.name} references {missing}, which no file in src/ps "
        f"defines. Rename the callers in the same commit as the function.")


def test_the_renamed_writability_helper_keeps_a_shim():
    """A stale caller on the share must get a message, not a parser error."""
    assert "function Assert-CgsWritable" in PS_UTILS, (
        "scanFileSystem.ps1 copies older than the rename call this. Keep the "
        "shim so they report the lock instead of dying on an unknown term.")
    block = re.search(r"function Assert-CgsWritable.*?\n}\n",
                      PS_UTILS, re.DOTALL).group(0)
    # Look at the code only; the help block above it names the replacements.
    code = block.split("#>", 1)[1]
    # Old behaviour: throws on failure, returns nothing on success. A stale
    # caller ignores return values, so it must not be given a fallback path.
    assert "Test-CgsWritable" in code
    assert "Resolve-CgsWritableTarget" not in code
    assert "throw" in code


def test_the_scanner_logs_which_cgsutils_it_loaded():
    assert "function Get-CgsUtilsBanner" in PS_UTILS
    assert "$script:CgsUtilsApi" in PS_UTILS
    assert "Write-CgsInfo (Get-CgsUtilsBanner)" in PS_SCANNER


# --- PowerShell -f binds tighter than + --------------------------------------

def _brokenFormatConcatenations(text):
    """Find `"...{0}..." + "..." -f $x`, where -f formats only the LAST piece.

    PowerShell's format operator binds tighter than string concatenation, so
    "a{0}" + "b" -f $x evaluates as "a{0}" + ("b" -f $x). Any placeholder in
    an earlier segment is emitted literally -- the user sees "{0}" instead of
    the path.
    """
    return [text[:m.start()].count("\n") + 1
            for m in re.finditer(r'\+\s*\r?\n\s*"[^"\n]*"\s+-f ', text)]


@pytest.mark.parametrize(
    "psPath",
    sorted((ROOT / "src" / "ps").glob("*.ps1")),
    ids=lambda p: p.name)
def test_powershell_format_operator_is_not_applied_to_a_concatenation(psPath):
    broken = _brokenFormatConcatenations(psPath.read_text())
    assert not broken, (
        f"{psPath.name}: line(s) {broken} apply -f to a concatenation, so "
        f"only the last string is formatted and earlier {{0}} placeholders "
        f"reach the user literally. Assign the message to a variable first.")
