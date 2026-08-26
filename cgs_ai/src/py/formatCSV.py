"""
=====================================================================
  Program Name  : formatCSV.py
  Author        : Manuel Figallo
  Purpose       : Convert a CSV into a styled Excel workbook with a SAS
                  ODS-style look and feel, suitable for distribution.
  Version       : 1.0beta
  Created       : 2026-08-26
  Last Modified : 2026-08-26

  Dependencies:
    openpyxl, imported lazily. Everything else is standard library. If
    openpyxl is absent the error explains how to install it.

  Description:
    The default "corporate" style is a navy title banner, blue header row,
    zebra striping and a frozen, auto-filtered header -- clean and
    business-appropriate, matching SAS ODS output.

  Input Parameters (required first):
    InputCsvPath    (REQUIRED, str) - source CSV.
    OutputExcelPath (REQUIRED, str) - destination .xlsx.
    FormatType      (optional, str, default "corporate") - corporate |
                      plain | minimal.
    SheetName       (optional, str, default "Report")
    Title           (optional, str) - banner text; defaults to the CSV name.
=====================================================================
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict

_HERE = Path(__file__).resolve()
if str(_HERE.parent.parent.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent.parent.parent))

from src.utils.helpers import ensureParent, readCsv  # noqa: E402
from src.utils.logger import logInfo                 # noqa: E402

__version__ = "1.0beta"

#: Palette per format type: (banner, header, stripe, header font colour).
FORMAT_STYLES: Dict[str, Dict[str, str]] = {
    "corporate": {"banner": "1F3864", "header": "2E75B6", "stripe": "DCE6F1",
                  "headerFont": "FFFFFF", "bannerFont": "FFFFFF"},
    "plain":     {"banner": "FFFFFF", "header": "D9D9D9", "stripe": "FFFFFF",
                  "headerFont": "000000", "bannerFont": "000000"},
    "minimal":   {"banner": "FFFFFF", "header": "FFFFFF", "stripe": "FFFFFF",
                  "headerFont": "000000", "bannerFont": "000000"},
}
DEFAULT_FORMAT_TYPE = "corporate"
MAX_COLUMN_WIDTH = 60


def formatCSV(InputCsvPath: str, OutputExcelPath: str,
              FormatType: str = DEFAULT_FORMAT_TYPE,
              SheetName: str = "Report",
              Title: str = "") -> Dict[str, Any]:
    """Render a CSV as a styled Excel workbook.

    Parameters:
        InputCsvPath (str)    - REQUIRED source CSV.
        OutputExcelPath (str) - REQUIRED destination .xlsx.
        FormatType (str)      - corporate (default) | plain | minimal.
                                "corporate" is a navy banner, blue header,
                                zebra striping.
        SheetName (str)       - worksheet name; default "Report".
        Title (str)           - banner text; defaults to the CSV filename.
    Returns:
        dict with OutputExcelPath, RowCount, ColumnCount and FormatType.
    Raises:
        ValueError  - a required parameter is missing or FormatType unknown.
        ImportError - openpyxl is not installed (message says how to fix).
        OSError     - the CSV cannot be read.

    Use in claims processing:
        Turn a raw scan or claims extract into a report an analyst or manager
        can open directly, without hand-formatting in Excel each cycle.
    """
    if not InputCsvPath or not str(InputCsvPath).strip():
        raise ValueError("required parameter 'InputCsvPath' is missing or empty")
    if not OutputExcelPath or not str(OutputExcelPath).strip():
        raise ValueError("required parameter 'OutputExcelPath' is missing or empty")
    if FormatType not in FORMAT_STYLES:
        raise ValueError(f"unknown FormatType {FormatType!r}; expected one of: "
                         f"{', '.join(FORMAT_STYLES)}")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "formatCSV requires openpyxl for Excel styling. Install it with "
            "'pip install openpyxl'.")

    rows = readCsv(InputCsvPath)
    columns = list(rows[0].keys()) if rows else []
    style = FORMAT_STYLES[FormatType]
    banner = Title or Path(InputCsvPath).stem

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SheetName

    # Row 1: title banner spanning the table width.
    sheet.cell(row=1, column=1, value=banner)
    cell = sheet.cell(row=1, column=1)
    cell.font = Font(bold=True, size=14, color=style["bannerFont"])
    cell.fill = PatternFill("solid", fgColor=style["banner"])
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26
    if columns:
        sheet.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=len(columns))

    # Row 2: column headers.
    for index, column in enumerate(columns, start=1):
        header = sheet.cell(row=2, column=index, value=column)
        header.font = Font(bold=True, color=style["headerFont"])
        header.fill = PatternFill("solid", fgColor=style["header"])
        header.alignment = Alignment(vertical="center", wrap_text=True)

    # Rows 3+: data with zebra striping.
    stripe = PatternFill("solid", fgColor=style["stripe"])
    for rowIndex, row in enumerate(rows, start=3):
        for colIndex, column in enumerate(columns, start=1):
            dataCell = sheet.cell(row=rowIndex, column=colIndex,
                                  value=row.get(column, ""))
            if FormatType == "corporate" and rowIndex % 2 == 1:
                dataCell.fill = stripe

    if columns:
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = (f"A2:{get_column_letter(len(columns))}"
                                 f"{max(2, len(rows) + 2)}")
        for index, column in enumerate(columns, start=1):
            widest = max([len(str(column))] +
                         [len(str(r.get(column, ""))) for r in rows[:200]] or [10])
            sheet.column_dimensions[get_column_letter(index)].width = \
                min(MAX_COLUMN_WIDTH, max(10, widest + 2))

    ensureParent(OutputExcelPath)
    workbook.save(OutputExcelPath)
    logInfo(f"formatted {len(rows)} row(s) x {len(columns)} column(s) "
            f"[{FormatType}] -> {OutputExcelPath}")
    return {"OutputExcelPath": OutputExcelPath, "RowCount": len(rows),
            "ColumnCount": len(columns), "FormatType": FormatType}
