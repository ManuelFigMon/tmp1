"""
=====================================================================
  Program Name  : scanFileSystem.py
  Author        : Manuel Figallo
  Purpose       : Scan one or more directory roots for keyword matches in
                  text files and emit one row per MATCH, with surrounding
                  context lines and token extraction. Optionally parses
                  structured performance metrics from log files via an
                  opt-in metric profile, which adds an Excel sheet.
  Version       : 1.0beta
  Created       : 2026-08-20
  Last Modified : 2026-08-26

  Dependencies:
    CSV output requires nothing beyond the Python standard library.
    XLSX output requires openpyxl, imported lazily and only when a metric
    profile is active or an .xlsx path is requested.

  Description:
    Runs unattended (Windows Task Scheduler, or the SAS SYSTASK wrappers in
    src/sas). Never prompts. Logs to stderr. Exits non-zero on failure so a
    scheduler can detect it.

    GRAIN (changed in 1.0beta): output is one row per KEYWORD MATCH, not one
    row per file. A file with three matches produces three rows; a file with
    no match produces none. This is what makes the output directly usable
    for claims-processing review, where the unit of interest is the matched
    line and its context.

  Input Parameters (required first):
    input_folder_root      (REQUIRED, list[str]) - root path(s) to search.
                             Accepts a list or a ';'-delimited string.
    extract_keyword        (REQUIRED, list[str]) - keywords to find. Without
                             at least one keyword there are no matches and
                             therefore no rows.
    output_file_path       (optional, str) - .csv or .xlsx by extension; a
                             directory auto-names scan_YYYYMMDD_HHMMSS.csv
                             inside it. Omitted entirely, writes
                             scan_YYYYMMDD_HHMMSS.csv to the current folder.
    file_extensions        (list[str], default ["log","txt","sas"])
    include_subdirectories (bool, default True)
    folder_exclusion_list  (list[str], default []) - nothing excluded unless set
    file_exclusion_list    (list[str], default []) - prefixes stripped from
                             the filename to derive the program name
    lines_above            (int, default 5)  - context lines captured BEFORE
    lines_below            (int, default 5)  - context lines captured AFTER
    nth_token_after        (int, default 1)  - which token after the keyword
    nth_token_before       (int, default 1)  - which token before the keyword
    numeric_token_after    (int, default 1)  - which NUMERIC token after
    date_from / date_to    (str, default None) - inclusive YYYY-MM-DD bounds
    date_field             (str, default "modified") - created/modified/accessed
    metric_profile         (str, default "none") - "none" | "sas_log".
                             When set, an EXCEL file is produced (the user is
                             told so in the log) with a second sheet of
                             structured metrics.

  Output (default columns, one row per match):
    SourceDir, FileName, Line, LinesAbove, LinesBelow, FullPath, LineNumber,
    Keyword, ExtractedString, NthTokenAfter, NthTokenBefore,
    NumericTokenAfter, LastToken, FirstToken, FileTimestamp, extension,
    file_size_bytes, created_time, modified_time, accessed_time, scanned_at

  Exit codes:
    0 = success, 2 = config error, 3 = I/O error.

  Change Log:
    v1.0beta - Output regrained to one row per keyword match with the column
               set above; configurable context window and token extraction;
               metric_profile now announces and produces Excel output.
=====================================================================
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

_HERE = Path(__file__).resolve()
if str(_HERE.parent.parent.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent.parent.parent))

from src.utils.helpers import (asList, ensureParent, fileTimes, isoNow,  # noqa: E402
                               timestampSuffix, writeCsv)
from src.utils.logger import logError, logInfo, logWarn  # noqa: E402

__version__ = "1.0beta"

# =====================================================================
# CONFIG -- defaults; every one has a matching CLI flag.
# =====================================================================

input_folder_root: List[str] = []
extract_keyword: List[str] = []
output_file_path: Optional[str] = None
file_extensions: List[str] = ["log", "txt", "sas"]
include_subdirectories: bool = True
folder_exclusion_list: List[str] = []
file_exclusion_list: List[str] = []
lines_above: int = 5
lines_below: int = 5
nth_token_after: int = 1
nth_token_before: int = 1
numeric_token_after: int = 1
date_from: Optional[str] = None
date_to: Optional[str] = None
date_field: str = "modified"
metric_profile: str = "none"

# =====================================================================
# Constants
# =====================================================================

VALID_DATE_FIELDS = ("created", "modified", "accessed")
DEFAULT_OUTPUT_PREFIX = "scan"

EXIT_OK, EXIT_CONFIG_ERROR, EXIT_IO_ERROR = 0, 2, 3

MATCH_SHEET = "Matches"
METRIC_SHEET = "Metrics"

#: The default output columns, in order. One row per keyword match.
MATCH_COLUMNS = [
    "SourceDir", "FileName", "Line", "LinesAbove", "LinesBelow", "FullPath",
    "LineNumber", "Keyword", "ExtractedString", "NthTokenAfter",
    "NthTokenBefore", "NumericTokenAfter", "LastToken", "FirstToken",
    "FileTimestamp", "extension", "file_size_bytes", "created_time",
    "modified_time", "accessed_time", "scanned_at",
]

#: Columns of the extra metric sheet produced when metric_profile is active.
METRIC_COLUMNS = [
    "FullPath", "ProgramName", "StepIndex", "StepLabel",
    "RealTimeSec", "CpuTimeSec",
]

#: Metric profile registry. Add an entry to support a new log format; no
#: change to the crawl or output code is required.
METRIC_PROFILES: Dict[str, Dict[str, Any]] = {
    "none": {"Active": False},
    "sas_log": {
        "Active": True,
        "StepPattern": r"^NOTE:\s+(?P<label>.+?)\s+used\s+\(Total process time\)",
        "Metrics": {
            "RealTimeSec": r"^\s*real time\s+(?P<value>[0-9:.]+)",
            # 'user cpu time' also matched so FULLSTIMER logs are not missed.
            "CpuTimeSec": r"^\s*(?:user\s+)?cpu time\s+(?P<value>[0-9:.]+)",
        },
        "Lookahead": 10,
    },
}

_NUMERIC = re.compile(r"^[+-]?\$?\d[\d,]*\.?\d*%?$")


# =====================================================================
# Token + duration helpers
# =====================================================================

def parseDuration(raw: str) -> Optional[float]:
    """Parse a duration string into seconds.

    Parameters:
        raw (str) - "0.05", "1.20 seconds", "1:03.05" (mm:ss) or
                    "1:00:30.00" (hh:mm:ss).
    Returns:
        float seconds, or None when unparseable.
    """
    if raw is None:
        return None
    text = str(raw).strip().lower().replace("seconds", "").replace("second", "").strip()
    if not text:
        return None
    match = re.match(r"^(\d+):(\d{1,2}):(\d{1,2}(?:\.\d+)?)$", text)
    if match:
        return int(match.group(1)) * 3600 + int(match.group(2)) * 60 + float(match.group(3))
    match = re.match(r"^(\d+):(\d{1,2}(?:\.\d+)?)$", text)
    if match:
        return int(match.group(1)) * 60 + float(match.group(2))
    match = re.match(r"^(\d+(?:\.\d+)?)$", text)
    return float(match.group(1)) if match else None


def isNumericToken(token: str) -> bool:
    """True when a token looks numeric (allows $ , . % and a sign).

    Parameters: token (str).
    Returns: bool.
    """
    return bool(_NUMERIC.match(token))


def extractTokens(line: str, keyword: str, nthAfter: int, nthBefore: int,
                  numericAfter: int) -> Dict[str, str]:
    """Pull the requested tokens out of the matched line.

    Parameters:
        line (str)         - the full matched line.
        keyword (str)      - the keyword that matched (case-insensitive).
        nthAfter (int)     - which whitespace token AFTER the keyword (1-based).
        nthBefore (int)    - which token BEFORE the keyword, counting backwards.
        numericAfter (int) - which NUMERIC token after the keyword (1-based).
    Returns:
        dict with NthTokenAfter, NthTokenBefore, NumericTokenAfter,
        FirstToken and LastToken. Missing tokens come back as ''.
    """
    tokens = line.split()
    lowered = [token.lower() for token in tokens]
    needle = keyword.lower()

    # The keyword may span several tokens ("real time"); anchor on the first.
    anchor = -1
    firstWord = needle.split()[0] if needle.split() else needle
    for index, token in enumerate(lowered):
        if firstWord in token:
            anchor = index
            break

    after = tokens[anchor + len(needle.split()):] if anchor >= 0 else []
    before = tokens[:anchor] if anchor > 0 else []
    numerics = [token for token in after if isNumericToken(token)]

    pick = lambda seq, n: seq[n - 1] if 0 < n <= len(seq) else ""
    return {
        "NthTokenAfter": pick(after, nthAfter),
        # counted backwards from the keyword: 1 = the token immediately before
        "NthTokenBefore": before[-nthBefore] if 0 < nthBefore <= len(before) else "",
        "NumericTokenAfter": pick(numerics, numericAfter),
        "FirstToken": tokens[0] if tokens else "",
        "LastToken": tokens[-1] if tokens else "",
    }


# =====================================================================
# Crawl / filter
# =====================================================================

def normalizePathText(text: str) -> str:
    """Lowercase, forward-slashed, trailing-slash-free form for comparison."""
    return str(text).replace("\\", "/").rstrip("/").lower()


def isFolderExcluded(path: Path, exclusions: Sequence[str]) -> bool:
    """True when an ancestor directory segment or a full-path prefix matches.

    Parameters:
        path (Path)                 - the candidate file.
        exclusions (sequence[str])  - folder names/tokens; empty excludes nothing.
    Returns:
        bool. Segment matching is exact, so excluding "Old" keeps "Older".
    """
    if not exclusions:
        return False
    segments = {part.lower() for part in path.parent.parts}
    normalized = normalizePathText(path)
    for raw in exclusions:
        token = str(raw).strip()
        if not token:
            continue
        if token.replace("\\", "/").rstrip("/").lower() in segments:
            return True
        prefix = normalizePathText(token)
        if prefix and (normalized == prefix or normalized.startswith(prefix + "/")):
            return True
    return False


def parseDateBoundary(raw: Optional[str], endOfDay: bool) -> Optional[dt.datetime]:
    """Parse an inclusive date bound.

    Parameters:
        raw (str)       - YYYY-MM-DD or ISO datetime; None/'' returns None.
        endOfDay (bool) - extend a bare date to 23:59:59.999999 (upper bound).
    Returns: datetime or None.
    Raises: ValueError on a malformed value.
    """
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip()
    try:
        value = dt.datetime.fromisoformat(text)
    except ValueError:
        raise ValueError(f"unparseable date {raw!r}; expected YYYY-MM-DD or ISO datetime")
    if endOfDay and len(text) == 10:
        value = value.replace(hour=23, minute=59, second=59, microsecond=999999)
    return value


def readTextLines(path: Path) -> List[str]:
    """Read a file as lines: UTF-8, then latin-1, then UTF-8 with replacement.

    Parameters: path (Path).
    Returns: list[str] without line terminators.
    """
    data = path.read_bytes()
    for encoding, errors in (("utf-8", "strict"), ("latin-1", "strict"), ("utf-8", "replace")):
        try:
            return data.decode(encoding, errors).splitlines()
        except (UnicodeDecodeError, LookupError):
            continue
    return []


def iterCandidateFiles(roots: Sequence[str], recurse: bool) -> Tuple[List[Path], int]:
    """List every file under the roots.

    Parameters:
        roots (sequence[str]) - directory roots; missing ones are logged, skipped.
        recurse (bool)        - walk subdirectories.
    Returns: (list of Paths, count of reachable roots).
    """
    found: List[Path] = []
    reachable = 0
    for rawRoot in roots:
        root = Path(str(rawRoot).strip())
        if not root.exists() or not root.is_dir():
            logWarn(f"root not found or not a directory, skipping: {root}")
            continue
        reachable += 1
        try:
            if recurse:
                for dirpath, _dirnames, filenames in os.walk(root, onerror=_walkError):
                    found.extend(Path(dirpath) / name for name in filenames)
            else:
                found.extend(child for child in root.iterdir() if child.is_file())
        except PermissionError as exc:
            logWarn(f"permission denied while walking {root}: {exc}")
    return found, reachable


def _walkError(exc: OSError) -> None:
    logWarn(f"cannot descend into {getattr(exc, 'filename', '?')}: {exc}")


# =====================================================================
# Metric profile
# =====================================================================

def parseMetricProfile(profile: Dict[str, Any], lines: Sequence[str],
                       fullPath: str, programName: str) -> List[Dict[str, Any]]:
    """Extract structured metric rows from a file's lines.

    Parameters:
        profile (dict)     - a METRIC_PROFILES entry.
        lines (sequence)   - the file's lines.
        fullPath (str)     - full path, carried onto every metric row.
        programName (str)  - filename stem, carried onto every metric row.
    Returns:
        list of dicts shaped by METRIC_COLUMNS; empty when inactive.
    """
    if not profile.get("Active"):
        return []
    stepPattern = re.compile(profile["StepPattern"], re.IGNORECASE)
    metricPatterns = {name: re.compile(pattern, re.IGNORECASE)
                      for name, pattern in profile["Metrics"].items()}
    lookahead = int(profile.get("Lookahead", 10))

    rows: List[Dict[str, Any]] = []
    for index, line in enumerate(lines):
        header = stepPattern.search(line)
        if not header:
            continue
        row: Dict[str, Any] = {
            "FullPath": fullPath,
            "ProgramName": programName,
            "StepIndex": len(rows) + 1,
            "StepLabel": (header.groupdict().get("label") or "").strip(),
        }
        window = lines[index + 1: index + 1 + lookahead]
        for name, pattern in metricPatterns.items():
            row[name] = None
            for candidate in window:
                hit = pattern.search(candidate)
                if hit:
                    row[name] = parseDuration(hit.groupdict().get("value", ""))
                    break
        rows.append(row)
    return rows


# =====================================================================
# Output
# =====================================================================

def defaultOutputName(directory: Optional[str] = None) -> str:
    """Build scan_YYYYMMDD_HHMMSS.csv, optionally inside `directory`."""
    name = f"{DEFAULT_OUTPUT_PREFIX}_{timestampSuffix()}.csv"
    return str(Path(directory) / name) if directory else name


def assertWritable(path: str) -> None:
    """Fail now, not after a long crawl, if the output cannot be written.

    Parameters: path (str) - the resolved output file.
    Returns: None.
    Raises: OSError naming the likely cause. The usual one is the workbook
            being open in Excel, which holds an exclusive lock on Windows.

    Without this the scan crawls every root -- minutes on a big share --
    and only then discovers it cannot write the answer.
    """
    target = Path(path)
    existed = target.exists()
    try:
        with open(target, "ab"):
            pass
    except OSError as exc:
        raise OSError(
            f"cannot write {target}: {exc.strerror or exc}. If the file is "
            f"open in Excel or another program, close it and run again, or "
            f"pass a different output_file_path.") from exc
    if not existed:
        try:
            target.unlink()
        except OSError:
            pass


def resolveOutputPath(raw: str) -> str:
    """Resolve the output target; a directory auto-names a timestamped CSV.

    Parameters: raw (str) - a file path, a directory, or ''.
    Returns: the resolved file path (parent directories created).
    """
    path = Path(str(raw).strip())
    looksLikeDir = (
        (path.exists() and path.is_dir())
        or str(raw).rstrip().endswith(("/", "\\"))
        or path.suffix.lower() not in (".csv", ".xlsx")
    )
    if looksLikeDir:
        path.mkdir(parents=True, exist_ok=True)
        generated = defaultOutputName(str(path))
        logInfo(f"output path is a directory; writing {generated}")
        return generated
    ensureParent(str(path))
    return str(path)


def writeExcel(matchRows: List[Dict[str, Any]], metricRows: List[Dict[str, Any]],
               outputPath: str) -> str:
    """Write matches (and metrics, when present) to an .xlsx workbook.

    Parameters:
        matchRows (list[dict])  - rows for the Matches sheet.
        metricRows (list[dict]) - rows for the Metrics sheet; omitted if empty.
        outputPath (str)        - .xlsx destination.
    Returns: the path written.
    Raises: ImportError when openpyxl is unavailable (message says how to fix).
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "Excel output requires openpyxl. Install it with "
            "'pip install openpyxl', or choose a .csv output path instead."
        )
    illegal = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
    clean = lambda v: illegal.sub("", v) if isinstance(v, str) else v

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(MATCH_SHEET)
    sheet.append(MATCH_COLUMNS)
    for row in matchRows:
        sheet.append([clean(row.get(column, "")) for column in MATCH_COLUMNS])
    if metricRows:
        metrics = workbook.create_sheet(METRIC_SHEET)
        metrics.append(METRIC_COLUMNS)
        for row in metricRows:
            metrics.append([clean(row.get(column, "")) for column in METRIC_COLUMNS])
    ensureParent(outputPath)
    workbook.save(outputPath)
    return outputPath


# =====================================================================
# Main entry point
# =====================================================================

def scanFileSystem(
    input_folder_root: Any,
    extract_keyword: Any,
    output_file_path: Optional[str] = None,
    file_extensions: Any = ("log", "txt", "sas"),
    include_subdirectories: bool = True,
    folder_exclusion_list: Any = (),
    file_exclusion_list: Any = (),
    lines_above: int = 5,
    lines_below: int = 5,
    nth_token_after: int = 1,
    nth_token_before: int = 1,
    numeric_token_after: int = 1,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    date_field: str = "modified",
    metric_profile: str = "none",
) -> Dict[str, Any]:
    """Scan directory roots for keyword matches; one result row per match.

    Parameters:
        input_folder_root      - REQUIRED root path(s); list or ';'-string.
        extract_keyword        - REQUIRED keyword(s); list or ';'-string.
        output_file_path       - .csv/.xlsx path, a directory, or None.
        file_extensions        - extensions to include (case-insensitive).
        include_subdirectories - recurse when True.
        folder_exclusion_list  - folder names/tokens to skip; [] skips nothing.
        file_exclusion_list    - prefixes stripped to derive the program name.
        lines_above            - context lines captured before the match.
        lines_below            - context lines captured after the match.
        nth_token_after        - which token after the keyword (1-based).
        nth_token_before       - which token before the keyword (1-based back).
        numeric_token_after    - which numeric token after the keyword.
        date_from / date_to    - inclusive YYYY-MM-DD bounds on date_field.
        date_field             - created | modified | accessed.
        metric_profile         - none | sas_log. When active, EXCEL output is
                                 produced with an extra Metrics sheet.
    Returns:
        dict with keys: matches (list[dict]), metrics (list[dict]),
        output (str path written), reachable (int roots reachable).
    Raises:
        ValueError - invalid metric_profile, date_field or date.

    Use in claims processing:
        Sweep SAS/ETL logs for terms such as "ERROR", ".accdb" or a claim
        number, and get back the exact line, its surrounding context and the
        adjacent tokens (for example the numeric count following a keyword)
        in a table an analyst can filter directly.
    """
    roots = asList(input_folder_root)
    keywords = asList(extract_keyword)
    extensions = {str(e).strip().lstrip(".").lower() for e in asList(file_extensions)}
    folderExclusions = asList(folder_exclusion_list)
    fileExclusions = asList(file_exclusion_list)

    if metric_profile not in METRIC_PROFILES:
        raise ValueError(f"unknown metric_profile {metric_profile!r}; "
                         f"expected one of: {', '.join(METRIC_PROFILES)}")
    if date_field not in VALID_DATE_FIELDS:
        raise ValueError(f"unknown date_field {date_field!r}; "
                         f"expected one of: {', '.join(VALID_DATE_FIELDS)}")
    dateLow = parseDateBoundary(date_from, endOfDay=False)
    dateHigh = parseDateBoundary(date_to, endOfDay=True)
    if dateLow and dateHigh and dateLow > dateHigh:
        raise ValueError(f"date_from ({date_from}) is after date_to ({date_to})")

    profile = METRIC_PROFILES[metric_profile]
    profileActive = bool(profile.get("Active"))

    # --- resolve the output target BEFORE crawling so a bad path fails fast ---
    if output_file_path and str(output_file_path).strip():
        target = resolveOutputPath(str(output_file_path))
    else:
        target = defaultOutputName()
        logInfo(f"output_file_path not supplied; writing {target}")

    # An active metric profile REQUIRES Excel, because the metrics go on a
    # second sheet. Tell the user plainly rather than silently switching.
    if profileActive and not target.lower().endswith(".xlsx"):
        original = target
        target = str(Path(target).with_suffix(".xlsx"))
        logInfo(f"metric_profile='{metric_profile}' is active, so EXCEL output "
                f"is produced: {target}")
        logInfo(f"(requested {Path(original).suffix or 'no extension'}; the extra "
                f"'{METRIC_SHEET}' sheet cannot be written to CSV)")

    # Prove the target is writable BEFORE the crawl: on a big share the
    # scan takes minutes, and losing it to a locked file is avoidable.
    assertWritable(target)

    scannedAt = isoNow()
    candidates, reachable = iterCandidateFiles(roots, include_subdirectories)
    logInfo(f"discovered {len(candidates)} file(s) across {reachable} reachable root(s)")

    matchRows: List[Dict[str, Any]] = []
    metricRows: List[Dict[str, Any]] = []

    for path in sorted(candidates, key=str):
        extension = path.suffix.lstrip(".").lower()
        if extensions and extension not in extensions:
            continue
        if isFolderExcluded(path, folderExclusions):
            continue

        try:
            stats = fileTimes(str(path))
        except OSError as exc:
            logWarn(f"cannot stat {path}: {exc}")
            continue

        times = {
            "created": dt.datetime.fromisoformat(stats["created_time"]),
            "modified": dt.datetime.fromisoformat(stats["modified_time"]),
            "accessed": dt.datetime.fromisoformat(stats["accessed_time"]),
        }
        if dateLow is not None and times[date_field] < dateLow:
            continue
        if dateHigh is not None and times[date_field] > dateHigh:
            continue

        try:
            lines = readTextLines(path)
        except OSError as exc:
            logWarn(f"cannot read {path}: {exc}")
            continue

        programName = path.stem
        for raw in fileExclusions:
            token = str(raw).strip()
            if token and programName.lower().startswith(token.lower()):
                programName = programName[len(token):]

        if profileActive:
            metricRows.extend(parseMetricProfile(profile, lines, str(path), programName))

        for lineIndex, line in enumerate(lines):
            lowered = line.lower()
            for keyword in keywords:
                if keyword.lower() not in lowered:
                    continue
                above = lines[max(0, lineIndex - lines_above):lineIndex]
                below = lines[lineIndex + 1: lineIndex + 1 + lines_below]
                row: Dict[str, Any] = {
                    "SourceDir": str(path.parent),
                    "FileName": path.name,
                    "Line": line,
                    "LinesAbove": "\n".join(above),
                    "LinesBelow": "\n".join(below),
                    "FullPath": str(path),
                    "LineNumber": lineIndex + 1,
                    "Keyword": keyword,
                    "ExtractedString": line,
                    "FileTimestamp": stats["modified_time"],
                    "extension": extension,
                    "file_size_bytes": stats["file_size_bytes"],
                    "created_time": stats["created_time"],
                    "modified_time": stats["modified_time"],
                    "accessed_time": stats["accessed_time"],
                    "scanned_at": scannedAt,
                }
                row.update(extractTokens(line, keyword, nth_token_after,
                                         nth_token_before, numeric_token_after))
                matchRows.append(row)

    if target.lower().endswith(".xlsx"):
        written = writeExcel(matchRows, metricRows, target)
        logInfo(f"wrote {len(matchRows)} match row(s) to sheet '{MATCH_SHEET}' in {written}")
        if metricRows:
            logInfo(f"wrote {len(metricRows)} metric row(s) to sheet '{METRIC_SHEET}'")
    else:
        written = writeCsv(matchRows, MATCH_COLUMNS, target)
        logInfo(f"wrote {len(matchRows)} match row(s) to {written}")

    return {"matches": matchRows, "metrics": metricRows,
            "output": written, "reachable": reachable}


def buildArgParser() -> argparse.ArgumentParser:
    """Build the command-line parser. Parameters: none."""
    parser = argparse.ArgumentParser(
        prog="scanFileSystem.py",
        description="Scan directory roots for keyword matches (one row per match).")
    parser.add_argument("--input-folder-root", nargs="+", default=None)
    parser.add_argument("--extract-keyword", nargs="+", default=None)
    parser.add_argument("--output-file-path", default=None)
    parser.add_argument("--file-extensions", nargs="+", default=None)
    parser.add_argument("--include-subdirectories", action=argparse.BooleanOptionalAction,
                        default=None)
    parser.add_argument("--folder-exclusion-list", nargs="*", default=None)
    parser.add_argument("--file-exclusion-list", nargs="*", default=None)
    parser.add_argument("--lines-above", type=int, default=5)
    parser.add_argument("--lines-below", type=int, default=5)
    parser.add_argument("--nth-token-after", type=int, default=1)
    parser.add_argument("--nth-token-before", type=int, default=1)
    parser.add_argument("--numeric-token-after", type=int, default=1)
    parser.add_argument("--date-from", default=None)
    parser.add_argument("--date-to", default=None)
    parser.add_argument("--date-field", default="modified", choices=list(VALID_DATE_FIELDS))
    parser.add_argument("--metric-profile", default="none", choices=list(METRIC_PROFILES))
    parser.add_argument("--version", action="version", version=f"scanFileSystem {__version__}")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    """CLI entry point. Parameters: argv (sequence[str]). Returns: exit code."""
    args = buildArgParser().parse_args(argv)
    roots = asList(args.input_folder_root if args.input_folder_root is not None
                   else input_folder_root)
    keywords = asList(args.extract_keyword if args.extract_keyword is not None
                      else extract_keyword)

    if not roots:
        logError("required parameter 'input_folder_root' is missing or empty; "
                 "pass --input-folder-root")
        return EXIT_CONFIG_ERROR
    if not keywords:
        logError("required parameter 'extract_keyword' is missing or empty; "
                 "pass --extract-keyword (no keywords means no matches)")
        return EXIT_CONFIG_ERROR

    logInfo(f"scanFileSystem {__version__} starting; profile={args.metric_profile}; "
            f"roots={len(roots)}; keywords={len(keywords)}")
    try:
        result = scanFileSystem(
            input_folder_root=roots,
            extract_keyword=keywords,
            output_file_path=args.output_file_path,
            file_extensions=(args.file_extensions if args.file_extensions is not None
                             else file_extensions),
            include_subdirectories=(include_subdirectories
                                    if args.include_subdirectories is None
                                    else args.include_subdirectories),
            folder_exclusion_list=args.folder_exclusion_list or [],
            file_exclusion_list=args.file_exclusion_list or [],
            lines_above=args.lines_above,
            lines_below=args.lines_below,
            nth_token_after=args.nth_token_after,
            nth_token_before=args.nth_token_before,
            numeric_token_after=args.numeric_token_after,
            date_from=args.date_from,
            date_to=args.date_to,
            date_field=args.date_field,
            metric_profile=args.metric_profile,
        )
    except ValueError as exc:
        logError(str(exc))
        return EXIT_CONFIG_ERROR
    except (OSError, ImportError) as exc:
        logError(f"cannot complete scan: {exc}")
        return EXIT_IO_ERROR

    if result["reachable"] == 0:
        logError("none of the supplied input_folder_root path(s) are reachable")
        return EXIT_IO_ERROR

    logInfo(f"done; {len(result['matches'])} match row(s), "
            f"{len(result['metrics'])} metric row(s)")
    return EXIT_OK


if __name__ == "__main__":
    sys.exit(main())
