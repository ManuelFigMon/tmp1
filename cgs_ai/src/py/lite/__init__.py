"""
=====================================================================
  Program Name  : __init__.py  (cgs_ai LITE)
  Author        : Manuel Figallo
  Purpose       : A single-file, drop-in build of cgs_ai carrying only the
                  two functions an end user needs to turn a CSV into a
                  formatted workbook and tell somebody it is ready.
  Version       : 1.0beta-lite
  Created       : 2026-08-28
  Last Modified : 2026-08-28

  Dependencies:
    STANDARD LIBRARY ONLY to load. openpyxl is imported lazily by formatCSV
    and is the single third-party requirement; sendEmail needs nothing but
    smtplib and email, both standard library.

  Description:
    The full package is a folder tree that has to be on sys.path. This file
    is deliberately ONE file with NO internal imports, so it can be executed
    straight off a UNC share:

        import runpy
        runpy.run_path(r"\\\\server\\...\\src\\py\\lite\\__init__.py",
                       run_name="__init__")
        from cgs_ai import formatCSV, sendEmail

    The second line works because running this file registers a module named
    "cgs_ai" in sys.modules (see REGISTRATION at the bottom). If the full
    cgs_ai package is already imported, this step is skipped and the full
    package wins -- the lite build never shadows it.

    An equally valid pattern, which does not rely on that registration:

        ns = runpy.run_path(PATH, run_name="cgs_ai_lite")
        formatCSV, sendEmail = ns["formatCSV"], ns["sendEmail"]

  Function Index:
    formatCSV  - CSV to a styled Excel workbook (SAS ODS look and feel)
    sendEmail  - SMTP notification to one or many recipients

  Change Log:
    v1.0beta-lite - First lite build. formatCSV and sendEmail only, with
                    every src.utils helper inlined so the file stands alone.
=====================================================================
"""

from __future__ import annotations

import csv
import smtplib
import sys
import types
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Dict, List, Optional

__version__ = "1.0beta-lite"
__all__ = ["formatCSV", "sendEmail", "__version__"]

# --------------------------------------------------------------------- #
# Inlined helpers. The full package imports these from src.utils; the lite
# build carries its own copies so the file has no internal dependencies.
# --------------------------------------------------------------------- #

#: Palette per format type: banner, header, stripe and font colours.
FORMAT_STYLES: Dict[str, Dict[str, str]] = {
    "corporate": {"banner": "1F3864", "header": "2E75B6", "stripe": "DCE6F1",
                  "headerFont": "FFFFFF", "bannerFont": "FFFFFF"},
    "plain":     {"banner": "FFFFFF", "header": "D9D9D9", "stripe": "FFFFFF",
                  "headerFont": "000000", "bannerFont": "000000"},
    "minimal":   {"banner": "FFFFFF", "header": "FFFFFF", "stripe": "FFFFFF",
                  "headerFont": "000000", "bannerFont": "000000"},
}
DEFAULT_FORMAT_TYPE = "corporate"
MAX_COLUMN_WIDTH = 60
DEFAULT_SMTP_SERVER = "smtp.example.com"
DEFAULT_SMTP_PORT = 25


def logInfo(message: str) -> None:
    """Write an informational line to stderr. Parameters: message (str)."""
    print(f"[cgs_ai-lite] INFO  {message}", file=sys.stderr)


def logError(message: str) -> None:
    """Write an error line to stderr. Parameters: message (str)."""
    print(f"[cgs_ai-lite] ERROR {message}", file=sys.stderr)


def asList(value: Any) -> List[str]:
    """Normalise a value into a list of non-empty strings.

    Parameters:
        value - a list/tuple, a ';'-delimited string, a scalar, or None.
    Returns: list of stripped, non-empty strings.
    """
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = [str(item) for item in value]
    else:
        items = str(value).split(";")
    return [item.strip() for item in items if str(item).strip()]


def ensureParent(path: str) -> Path:
    """Create the parent directory of `path` if needed. Returns: the Path."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def readCsv(inputPath: str) -> List[Dict[str, str]]:
    """Read a CSV into a list of dicts.

    Parameters: inputPath (str) - the CSV to read.
    Returns: list of row dicts keyed by column name.
    Raises: OSError if the file cannot be read.
    """
    with open(inputPath, encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


# --------------------------------------------------------------------- #
# formatCSV
# --------------------------------------------------------------------- #

def formatCSV(InputCsvPath: str, OutputExcelPath: str,
              FormatType: str = DEFAULT_FORMAT_TYPE,
              SheetName: str = "Report",
              Title: str = "") -> Dict[str, Any]:
    """Render a CSV as a styled Excel workbook.

    Parameters:
        InputCsvPath (str)    - REQUIRED source CSV.
        OutputExcelPath (str) - REQUIRED destination .xlsx.
        FormatType (str)      - corporate (default) | plain | minimal.
                                "corporate" is a navy banner, blue header,
                                zebra striping.
        SheetName (str)       - worksheet name; default "Report".
        Title (str)           - banner text; defaults to the CSV filename.
    Returns:
        dict with OutputExcelPath, RowCount, ColumnCount and FormatType.
    Raises:
        ValueError  - a required parameter is missing or FormatType unknown.
        ImportError - openpyxl is not installed (message says how to fix).
        OSError     - the CSV cannot be read.

    Use in claims processing:
        Turn a claims extract into a report an analyst or manager can open
        directly, without hand-formatting it in Excel every cycle.
    """
    if not InputCsvPath or not str(InputCsvPath).strip():
        raise ValueError("required parameter 'InputCsvPath' is missing or empty")
    if not OutputExcelPath or not str(OutputExcelPath).strip():
        raise ValueError("required parameter 'OutputExcelPath' is missing or empty")
    if FormatType not in FORMAT_STYLES:
        raise ValueError(f"unknown FormatType {FormatType!r}; expected one of: "
                         f"{', '.join(FORMAT_STYLES)}")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "formatCSV requires openpyxl for Excel styling. Install it with "
            "'pip install openpyxl'.")

    rows = readCsv(InputCsvPath)
    columns = list(rows[0].keys()) if rows else []
    style = FORMAT_STYLES[FormatType]
    banner = Title or Path(InputCsvPath).stem

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SheetName

    # Row 1: title banner spanning the table width.
    cell = sheet.cell(row=1, column=1, value=banner)
    cell.font = Font(bold=True, size=14, color=style["bannerFont"])
    cell.fill = PatternFill("solid", fgColor=style["banner"])
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26
    if columns:
        sheet.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=len(columns))

    # Row 2: column headers.
    for index, column in enumerate(columns, start=1):
        header = sheet.cell(row=2, column=index, value=column)
        header.font = Font(bold=True, color=style["headerFont"])
        header.fill = PatternFill("solid", fgColor=style["header"])
        header.alignment = Alignment(vertical="center", wrap_text=True)

    # Rows 3+: data with zebra striping.
    stripe = PatternFill("solid", fgColor=style["stripe"])
    for rowIndex, row in enumerate(rows, start=3):
        for colIndex, column in enumerate(columns, start=1):
            dataCell = sheet.cell(row=rowIndex, column=colIndex,
                                  value=row.get(column, ""))
            if FormatType == "corporate" and rowIndex % 2 == 1:
                dataCell.fill = stripe

    if columns:
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = (f"A2:{get_column_letter(len(columns))}"
                                 f"{max(2, len(rows) + 2)}")
        for index, column in enumerate(columns, start=1):
            widest = max([len(str(column))] +
                         [len(str(r.get(column, ""))) for r in rows[:200]] or [10])
            sheet.column_dimensions[get_column_letter(index)].width = \
                min(MAX_COLUMN_WIDTH, max(10, widest + 2))

    ensureParent(OutputExcelPath)
    workbook.save(OutputExcelPath)
    logInfo(f"formatted {len(rows)} row(s) x {len(columns)} column(s) "
            f"[{FormatType}] -> {OutputExcelPath}")
    return {"OutputExcelPath": OutputExcelPath, "RowCount": len(rows),
            "ColumnCount": len(columns), "FormatType": FormatType}


# --------------------------------------------------------------------- #
# sendEmail
# --------------------------------------------------------------------- #

def sendEmail(To: Any, From: str, Subject: str, Body: str,
              SmtpServer: Optional[str] = None, Port: Optional[int] = None,
              Html: bool = False, Attachments: Any = ()) -> Dict[str, Any]:
    """Send an email alert over SMTP.

    Parameters:
        To (list|str)     - REQUIRED recipient(s); ';'-delimited string allowed.
        From (str)        - REQUIRED sender address.
        Subject (str)     - REQUIRED subject line.
        Body (str)        - REQUIRED message body.
        SmtpServer (str)  - SMTP host; defaults to smtp.example.com.
        Port (int)        - SMTP port; defaults to 25.
        Html (bool)       - send the body as HTML instead of plain text.
        Attachments       - optional file path(s) to attach.
    Returns:
        dict with To (list), Subject, SmtpServer, Port and Sent (bool).
    Raises:
        ValueError            - a required parameter is missing.
        smtplib.SMTPException - the server rejected the message.
        OSError               - the server is unreachable.

    Use in claims processing:
        Notify the operations mailbox when a claims report is ready,
        including the row count and the output location, so nobody has to
        watch the job.

    NOTE: the lite build has no .env loader, so SmtpServer and Port are NOT
    read from configuration -- pass them explicitly, or you get the
    smtp.example.com default, which will not deliver anything.
    """
    recipients: List[str] = asList(To)
    if not recipients:
        raise ValueError("required parameter 'To' is missing or empty")
    for name, value in (("From", From), ("Subject", Subject), ("Body", Body)):
        if value is None or str(value).strip() == "":
            raise ValueError(f"required parameter '{name}' is missing or empty")

    server = SmtpServer or DEFAULT_SMTP_SERVER
    port = int(Port or DEFAULT_SMTP_PORT)

    message = EmailMessage()
    message["To"] = ", ".join(recipients)
    message["From"] = From
    message["Subject"] = Subject
    message.set_content(Body, subtype="html" if Html else "plain")

    for attachment in asList(Attachments):
        path = Path(attachment)
        if not path.is_file():
            logError(f"attachment not found, skipping: {path}")
            continue
        message.add_attachment(path.read_bytes(), maintype="application",
                               subtype="octet-stream", filename=path.name)

    logInfo(f"sending mail to {len(recipients)} recipient(s) via {server}:{port}")
    with smtplib.SMTP(server, port, timeout=30) as smtp:
        smtp.send_message(message)
    logInfo(f"sent: {Subject}")
    return {"To": recipients, "Subject": Subject, "SmtpServer": server,
            "Port": port, "Sent": True}


# --------------------------------------------------------------------- #
# REGISTRATION
#
# runpy.run_path() executes this file but does not create an importable
# module, so `from cgs_ai import formatCSV` would fail on the next line.
# Publishing a module object into sys.modules makes that import work.
#
# The guard matters: if the FULL cgs_ai package is already imported, it is
# left alone. The lite build never shadows the real one.
# --------------------------------------------------------------------- #

def _register(moduleName: str = "cgs_ai") -> None:
    """Publish this file's functions as an importable module.

    Parameters:
        moduleName (str) - the name to register under; default "cgs_ai".
    Returns: None. Does nothing if that name is already imported.
    """
    if moduleName in sys.modules:
        return
    module = types.ModuleType(moduleName)
    module.__doc__ = f"cgs_ai lite build {__version__} (formatCSV, sendEmail)"
    module.__version__ = __version__
    module.__all__ = list(__all__)
    module.formatCSV = formatCSV
    module.sendEmail = sendEmail
    sys.modules[moduleName] = module


_register()
