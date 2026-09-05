"""
=====================================================================
  Program Name  : copyExcelSheet2CSV.py
  Author        : Manuel Figallo
  Purpose       : Export one worksheet of an Excel workbook to CSV, refusing
                  to proceed when the sheet is not shaped for flat output.
  Version       : 1.0beta
  Created       : 2026-08-26
  Last Modified : 2026-08-26

  Dependencies:
    openpyxl, imported lazily.

  Description:
    Validates the sheet BEFORE writing anything and stops on the first
    problem, rather than emitting a malformed CSV that fails downstream.
    Validation covers: sheet missing, sheet empty, a blank/duplicate header,
    and merged cells across the header row (which silently drop data).

  Input Parameters (required first):
    InputExcelPath (REQUIRED, str) - source .xlsx.
    SheetName      (REQUIRED, str) - worksheet to export.
    OutputCsvPath  (REQUIRED, str) - destination .csv.
    HeaderRow      (optional, int, default 1) - 1-based row holding the
                     column names. Use 2 for a workbook with a title banner
                     on row 1 -- which is exactly what formatCSV produces,
                     so formatCSV output round-trips with HeaderRow=2.
=====================================================================
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List

_HERE = Path(__file__).resolve()
if str(_HERE.parent.parent.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent.parent.parent))

from src.utils.helpers import ensureParent, writeCsv  # noqa: E402
from src.utils.logger import logError, logInfo        # noqa: E402

__version__ = "1.0beta"


class SheetNotCsvReadyError(ValueError):
    """Raised when a worksheet cannot be represented as flat CSV."""


def copyExcelSheet2CSV(InputExcelPath: str, SheetName: str,
                       OutputCsvPath: str,
                       HeaderRow: int = 1) -> Dict[str, Any]:
    """Export one Excel worksheet to CSV, validating it first.

    Parameters:
        InputExcelPath (str) - REQUIRED path to the .xlsx workbook.
        SheetName (str)      - REQUIRED worksheet name to export.
        OutputCsvPath (str)  - REQUIRED destination .csv path.
        HeaderRow (int)      - 1-based row containing the column names;
                               default 1. Pass 2 for a sheet with a title
                               banner on row 1 (formatCSV output).
    Returns:
        dict with OutputCsvPath, RowCount, ColumnCount and Columns.
    Raises:
        ValueError              - a required parameter is missing.
        SheetNotCsvReadyError   - the sheet is not shaped for CSV: it does
                                  not exist, is empty, has blank or duplicate
                                  header names, or has merged cells crossing
                                  the header row. Nothing is written.
        ImportError             - openpyxl is not installed.

    Use in claims processing:
        Convert a hand-maintained Excel reference workbook (fee schedules,
        denial-code mappings) into the CSV a pipeline can consume, while
        catching formatting mistakes at the source instead of downstream.
    """
    for name, value in (("InputExcelPath", InputExcelPath),
                        ("SheetName", SheetName),
                        ("OutputCsvPath", OutputCsvPath)):
        if not value or not str(value).strip():
            raise ValueError(f"required parameter '{name}' is missing or empty")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "copyExcelSheet2CSV requires openpyxl. Install it with "
            "'pip install openpyxl'.")

    workbook = load_workbook(InputExcelPath, data_only=True)
    if SheetName not in workbook.sheetnames:
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' not found in {InputExcelPath}. "
            f"Available sheets: {', '.join(workbook.sheetnames)}")
    sheet = workbook[SheetName]

    # --- validate before writing anything -----------------------------------
    if HeaderRow < 1:
        raise ValueError("HeaderRow must be 1 or greater")
    merged = [str(r) for r in getattr(sheet, "merged_cells", []).ranges] \
        if getattr(sheet, "merged_cells", None) else []
    # Only merges ON the header row break CSV export; a merged title banner
    # ABOVE the header is fine and is simply skipped.
    headerMerges = [r for r in merged if sheet[r.split(":")[0]].row == HeaderRow]
    if headerMerges:
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' has merged cells across the header row "
            f"({', '.join(headerMerges)}). Un-merge them: merged headers cannot "
            f"be represented in CSV and would silently drop columns.")

    allRows: List[List[Any]] = [list(r) for r in sheet.iter_rows(values_only=True)]
    rows = allRows[HeaderRow - 1:]          # discard anything above the header
    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' is empty; nothing to export.")

    header = ["" if c is None else str(c).strip() for c in rows[0]]
    while header and header[-1] == "":
        header.pop()
    if not header:
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' has no usable header row.")
    if any(name == "" for name in header):
        blanks = [i + 1 for i, n in enumerate(header) if n == ""]
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' has blank header name(s) in column(s) "
            f"{blanks}. Every column needs a name for CSV output.")
    duplicates = {n for n in header if header.count(n) > 1}
    if duplicates:
        raise SheetNotCsvReadyError(
            f"worksheet '{SheetName}' has duplicate header name(s): "
            f"{', '.join(sorted(duplicates))}. Column names must be unique.")

    # --- validated: write ----------------------------------------------------
    records = []
    for raw in rows[1:]:
        padded = list(raw) + [None] * (len(header) - len(raw))
        records.append({name: ("" if padded[i] is None else padded[i])
                        for i, name in enumerate(header)})

    ensureParent(OutputCsvPath)
    written = writeCsv(records, header, OutputCsvPath)
    logInfo(f"exported '{SheetName}': {len(records)} row(s) x {len(header)} "
            f"column(s) -> {written}")
    return {"OutputCsvPath": written, "RowCount": len(records),
            "ColumnCount": len(header), "Columns": header}
